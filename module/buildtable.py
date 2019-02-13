# import copy
from copy import deepcopy
from openpyxl.utils import get_column_letter, column_index_from_string


# hex操作基类
class HexBase(object):
	"""操作hex"""
	def build_hex_data(self, dataList):
		"""创建hex有效数据, dataList:二维列表"""
		self.hexDataList = []
		# 创建中断MO初始化表hex
		for subList in dataList:
			# 根据结构体中数据长度类型分别进行处理
			for index in range(len(subList)):
				if self.structType[index] == "uint8":
					# 根据十六进制数据或十进制数据分别处理
					if subList[index][0:2] == "0x" or subList[index][0:2] == "0X":
						self.hexDataList.append("{0:02X}".format(int(subList[index][0:-1], 16)))
					else:
						self.hexDataList.append("{0:02X}".format(int(subList[index][0:-1])))
				elif self.structType[index] == "uint16":
					if subList[index][0:2] == "0x" or subList[index][0:2] == "0X":
						tmpHexStr = "{0:04X}".format(int(subList[index][0:-1], 16))
					else:
						tmpHexStr = "{0:04X}".format(int(subList[index][0:-1]))
					self.hexDataList.append(tmpHexStr[2:4])
					self.hexDataList.append(tmpHexStr[0:2])
				elif self.structType[index] == "uint32":
					if subList[index][0:2] == "0x" or subList[index][0:2] == "0X":
						tmpHexStr = "{0:08X}".format(int(subList[index][0:-1], 16))
					else:
						tmpHexStr = "{0:08X}".format(int(subList[index][0:-1]))
						tmpHexStr = "{0:08X}".format(int(subList[index][0:-1]))
					self.hexDataList.append(tmpHexStr[6:8])
					self.hexDataList.append(tmpHexStr[4:6])
					self.hexDataList.append(tmpHexStr[2:4])
					self.hexDataList.append(tmpHexStr[0:2])
		# print(self.hexDataList)

	def build_table_len_hex_data(self, lenData):
		"""创建表长度hex数据"""
		# 创建表长度hex
		if self.lenType == "uint8":
			self.lenHexDataList.append("{0:02X}".format(int(lenData)))
		elif self.lenType == "uint16":
			tmpHexStr = "{0:04X}".format(int(lenData))
			self.lenHexDataList.append(tmpHexStr[2:4])
			self.lenHexDataList.append(tmpHexStr[0:2])
		elif self.lenType == "uint32":
			tmpHexStr = "{0:08X}".format(int(lenData))
			self.lenHexDataList.append(tmpHexStr[6:8])
			self.lenHexDataList.append(tmpHexStr[4:6])
			self.lenHexDataList.append(tmpHexStr[2:4])
			self.lenHexDataList.append(tmpHexStr[0:2])
		# print(self.lenHexDataList)

	def modify_hex_data(self, originalhexDataList, addr, dataLen, hexDataList):
		"""修改原始hex数据"""
		# 如果源HEX数据为空，默认未选择hex，以下不执行
		if originalhexDataList != []:
			if int(addr, 16) >= int("0x00c40000", 16):
				# block = 4
				self.offset = ((int(addr, 16) - int("0x00c10000", 16) + int("0x0000F000", 16))//32) + 5
			elif int(addr, 16) >= int("0x00c30000", 16):
				# block = 3
				self.offset = ((int(addr, 16) - int("0x00c10000", 16) + int("0x0000F000", 16))//32) + 4
			elif int(addr, 16) >= int("0x00c20000", 16):
				# block = 2
				self.offset = ((int(addr, 16) - int("0x00c10000", 16) + int("0x0000F000", 16))//32) + 3
			elif int(addr, 16) >= int("0x00c10000", 16):
				# block = 1
				self.offset = ((int(addr, 16) - int("0x00c10000", 16) + int("0x0000F000", 16))//32) + 2
			else:
				# block = 0
				self.offset = ((int(addr, 16) - int("0x00c00000", 16))//32) + 1

			self.byteOffset = int(addr, 16)%32
			# print(self.offset)
			# print(self.byteOffset)
			# print(originalhexDataList)

			offset = deepcopy(self.offset)
			byteOffset = deepcopy(self.byteOffset)
			# 清零				
			for index in range(dataLen):
				originalhexDataList[offset] = originalhexDataList[offset][0:9] + originalhexDataList[offset][9:9+byteOffset*2] + \
											"00" + originalhexDataList[offset][9+byteOffset*2 + 2:]
				byteOffset += 1
				if byteOffset > 31:
					byteOffset = 0
					if len(originalhexDataList[offset+1]) < 75: # 75代表每一数据行的字符个数
						offset += 2
					else:
						offset += 1

			offset = deepcopy(self.offset)
			byteOffset = deepcopy(self.byteOffset)
			# 重新写入
			for index in range(len(hexDataList)):
				originalhexDataList[offset] = originalhexDataList[offset][0:9] + originalhexDataList[offset][9:9+byteOffset*2] + \
											hexDataList[index] + originalhexDataList[offset][9+byteOffset*2 + 2:]
				byteOffset += 1
				if byteOffset > 31:
					byteOffset = 0
					if len(originalhexDataList[offset+1]) < 75: # 75代表每一数据行的字符个数
						offset += 2
					else:
						offset += 1


# 中断MO初始化表
class CanFullIdNameISR(HexBase):
	"""CAN_FULL_ID_NAME_ISR"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.msgValidDataListISR = []
		self.msgDesChListISR = []
		self.signalValidDataListISR = []
		self.signalDesChListISR = []
		self.CanFullIDNameISRList = []
		self.CAN_FULL_ID_NAME_ISR = []		
		self.structType = ["uint32", "uint8", "uint8", "uint16", "uint32", "uint8", "uint8", "uint8", "uint8", "uint16"]
		self.structLen = 18
		# self.tableLenAddr = "0x00c0a59f"
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["CAN_FULL_ID_NAME_ISR"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["CAN_FULL_ID_NAME_ISR"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c0b5f6"
		# self.tableLen = 60
		self.tableAddr = self.config.addrInfo["CAN_FULL_ID_NAME_ISR"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["CAN_FULL_ID_NAME_ISR"]["tableLen"])
		self.hexDataList = []

	def get_valid_data(self, msgRoute, signalRoute):
		"""获取有效数据"""
		# 深拷贝一份路由数据，不破坏原有数据
		msgDataList = deepcopy(msgRoute.dataList)
		msgHeaderList = deepcopy(msgRoute.headerList)
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)
		# 轮询普通报文数据
		for subList in msgDataList:
			# 判断子列表是否为中断报文数据
			# if subList[column_index_from_string('M') - 1] == '1':
			if subList[msgHeaderList.index("RxInterrupt")] == '1':
				# 每循环一次深拷贝一份报文表头，因为每次循环都对原始表头修改
				self.msgValidHeaderList = deepcopy(msgHeaderList)				
				# 弹出序号和目标通道，同一发送通道相同ID发往不同通道的余下内容相同
				# num = subList.pop(column_index_from_string('A') - 1)
				# MsgDesCh = subList.pop(column_index_from_string('F') - 2)
				num = subList.pop(self.msgValidHeaderList.index("LineNumber"))
				self.msgValidHeaderList.pop(self.msgValidHeaderList.index("LineNumber"))
				# print(self.msgValidHeaderList)
				MsgDesCh = subList.pop(self.msgValidHeaderList.index("TxChannle"))
				self.msgValidHeaderList.pop(self.msgValidHeaderList.index("TxChannle"))
				# print(self.msgValidHeaderList)
				# 如果子列表数据首次出现，则保存，否则只将相应目的通道附加至目的通道列表
				if subList not in self.msgValidDataListISR:
					self.msgValidDataListISR.append(subList)
					self.msgDesChListISR.append([subList])
					if MsgDesCh != '0':
						self.msgDesChListISR[len(self.msgValidDataListISR) - 1].append(MsgDesCh)
				else:
					if MsgDesCh != '0':
						self.msgDesChListISR[len(self.msgValidDataListISR) - 1].append(MsgDesCh)
		# print(self.msgValidDataListISR)
		# print(self.msgDesChListISR)

		# 对普通报文信息和报文目标通道按ID大小进行排序,再以通道大小排序
		self.msgValidDataListISR = sorted(self.msgValidDataListISR, key=lambda subList:[int(subList[self.msgValidHeaderList.index("TxCANID")], 16), \
																						int(subList[self.msgValidHeaderList.index("RxChannel")])])
		self.msgDesChListISR = sorted(self.msgDesChListISR, key=lambda subList:[int(subList[0][self.msgValidHeaderList.index("TxCANID")], 16), \
																				int(subList[0][self.msgValidHeaderList.index("RxChannel")])])
		# print(self.msgValidDataListISR)
		# print(self.msgDesChListISR)
		# 轮询信号报文数据
		for subList in signalDataList:
			pass

	def __msgDataHandlingISR1(self, index):
		"""中断初始化表接收部分"""
		retList = []
		# retList.append(self.msgValidDataListISR[index][column_index_from_string('C') - 2] + 'u')
		# retList.append(self.msgValidDataListISR[index][column_index_from_string('K') - 3] + 'u')
		retList.append(self.msgValidDataListISR[index][self.msgValidHeaderList.index("TxCANID")] + 'u')
		retList.append(self.msgValidDataListISR[index][self.msgValidHeaderList.index("RxChannel")] + 'u')
		retList.append('1u')
		retList.append(str(self.MO_num) + 'u')	
		# retList.append(self.msgValidDataListISR[index][column_index_from_string('L') - 3] + 'u')
		retList.append(self.msgValidDataListISR[index][self.msgValidHeaderList.index("RxMsk")] + 'u')
		retList.append('0u')
		# retList.append(self.msgValidDataListISR[index][column_index_from_string('J') - 3] + 'u')
		retList.append(self.msgValidDataListISR[index][self.msgValidHeaderList.index("RxDLC")] + 'u')
		retList.append('1u')
		retList.append('0xffu')
		retList.append('0u')
		self.MO_num += 1

		# print(retList)
		return retList

	def __msgDataHandlingISR2(self, i, des_ch):
		"""中断初始化表发送部分"""
		retList = []	
		# retList.append(self.msgDesChListISR[i][0][column_index_from_string('C') - 2] + 'u')
		retList.append(self.msgDesChListISR[i][0][self.msgValidHeaderList.index("TxCANID")] + 'u')
		retList.append(str(des_ch) + 'u')
		retList.append('0u')
		retList.append(str(self.MO_num) + 'u')		
		# retList.append(self.msgDesChListISR[i][0][column_index_from_string('L') - 3] + 'u')
		retList.append(self.msgDesChListISR[i][0][self.msgValidHeaderList.index("RxMsk")] + 'u')
		retList.append('0u')
		# retList.append(self.msgDesChListISR[i][0][column_index_from_string('J') - 3] + 'u')
		retList.append(self.msgDesChListISR[i][0][self.msgValidHeaderList.index("RxDLC")] + 'u')
		retList.append('0u')
		retList.append("0xffu")
		retList.append('0u')
		self.MO_num += 1

		# print(retList)
		return retList

	def data_handle(self):
		"""数据处理"""
		# 用于标记中断MO序号
		self.MO_num = 0
		# 轮询中断接收处理报文
		for index in range(len(self.msgValidDataListISR)):
			self.CanFullIDNameISRList.append(self.__msgDataHandlingISR1(index))
		# print(self.CanFullIDNameISRList)

		# 轮询中断发送处理报文
		for i in range(len(self.msgDesChListISR)):
			# 对发送通道按ID大小进行排序
			MsgDesCh_ISR = sorted(self.msgDesChListISR[i][1:])
			# 轮询所有目的通道
			for j in range(len(self.msgDesChListISR[i]) - 1):
				des_ch = MsgDesCh_ISR[j]
				self.CanFullIDNameISRList.append(self.__msgDataHandlingISR2(i, des_ch))
		# print(self.CanFullIDNameISRList)

	def build_table(self):
		"""创建数组表"""
		self.CAN_FULL_ID_NAME_ISR.append("const uint8 Can_Isr_Tab_Len __at(0x00c0a59f) = " + str(len(self.CanFullIDNameISRList)) + "u;\n")
		self.CAN_FULL_ID_NAME_ISR.append("\n")
		self.CAN_FULL_ID_NAME_ISR.append("/* ================================ interrupt msg ================================== */\n")
		self.CAN_FULL_ID_NAME_ISR.append("const SHUGE MULTICAN_FULL_ID CAN_FULL_ID_NAME_ISR[CAN_CFG_INT_LEN] __at(0x00c0b5f6)\n")
		self.CAN_FULL_ID_NAME_ISR.append("=\n")
		self.CAN_FULL_ID_NAME_ISR.append("{\n")
		for index, subList in enumerate(self.CanFullIDNameISRList):
			self.CAN_FULL_ID_NAME_ISR.append('\t{')
			self.CAN_FULL_ID_NAME_ISR.append(','.join(subList))
			self.CAN_FULL_ID_NAME_ISR.append('},' + "/*" + str(index) + "*/" + '\n')
		else:
			# 如果列表为空，则以0，填充数组，以防工程编译时不通过.
			if not self.CanFullIDNameISRList:
				self.CAN_FULL_ID_NAME_ISR.append('\t0,\n')
		self.CAN_FULL_ID_NAME_ISR.append('};')


# 中断轮询表、报文轮询表和索引表的基类
class RoutingTable(object):
	# 获取有效数据
	def get_valid_data(self, msgRoute, signalRoute):
		"""获取有效数据"""
		# 深拷贝一份路由数据，不破坏原有数据
		msgDataList = deepcopy(msgRoute.dataList)
		msgHeaderList = deepcopy(msgRoute.headerList)
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)

		if msgDataList:
			for subList in msgDataList:
				# 每循环一次深拷贝一份报文表头，因为每次循环都对原始表头修改
				self.msgValidHeaderList = deepcopy(msgHeaderList)	
				# 弹出序号和目标通道，同一发送通道相同ID发往不同通道的余下内容相同
				# num = subList.pop(column_index_from_string('A') - 1)
				# MsgDesCh = subList.pop(column_index_from_string('F') - 2)
				num = subList.pop(self.msgValidHeaderList.index("LineNumber"))
				self.msgValidHeaderList.pop(self.msgValidHeaderList.index("LineNumber"))
				MsgDesCh = subList.pop(self.msgValidHeaderList.index("TxChannle"))
				self.msgValidHeaderList.pop(self.msgValidHeaderList.index("TxChannle"))

				if subList not in self.msgValidDataList:
					self.msgValidDataList.append(subList)
					self.msgDesChList.append([subList[self.msgValidHeaderList.index("TxCANID")]])
					if MsgDesCh != '0':
						self.msgDesChList[len(self.msgValidDataList) - 1].append(MsgDesCh)
					if subList[self.msgValidHeaderList.index("RxInterrupt")] == "0":
						pass
						# self.msgDesChList[len(self.msgValidDataList) - 1].append(MsgDesCh)
					else:
						self.msgValidDataListISR.append(subList)
						self.msgDesChListISR.append([subList])
						if MsgDesCh != '0':
							self.msgDesChListISR[len(self.msgValidDataListISR) - 1].append(MsgDesCh)

				else:
					if MsgDesCh != '0':
						self.msgDesChList[self.msgValidDataList.index(subList)].append(MsgDesCh)
					if subList[self.msgValidHeaderList.index("RxInterrupt")] == "0":
						pass
						# self.msgDesChList[self.msgValidDataList.index(subList)].append(MsgDesCh)
					else:
						if MsgDesCh != '0':
							self.msgDesChListISR[len(self.msgValidDataListISR) - 1].append(MsgDesCh)
			# print(self.msgValidDataList)
			# print(self.msgDesChList)
			# print(len(self.msgValidDataList))
			# print(len(self.msgDesChList))
			# 将中断有效数据和对应的报文通道先以ID大小排序，再以通道大小排序
			self.msgValidDataListISR = sorted(self.msgValidDataListISR, key=lambda subList:[int(subList[self.msgValidHeaderList.index("TxCANID")], 16), \
																							int(subList[self.msgValidHeaderList.index("RxChannel")])])
			self.msgDesChListISR = sorted(self.msgDesChListISR, key=lambda subList:[int(subList[0][self.msgValidHeaderList.index("TxCANID")], 16), \
																					int(subList[0][self.msgValidHeaderList.index("RxChannel")])])
			# for i in range(len(self.msgDesChListISR)):
			# 	self.msgDesChListISR[]
			# print(self.msgValidDataListISR)
			# print(self.msgDesChListISR)

		# 提取信号有效头列表
		if signalDataList:
			self.sigValidHeaderList = [signalHeaderList[signalHeaderList.index("RxCANID")],\
									   signalHeaderList[signalHeaderList.index("RxChannel")],\
									   signalHeaderList[signalHeaderList.index("RxDTC")],\
									   signalHeaderList[signalHeaderList.index("RxPeriod")],\
									   signalHeaderList[signalHeaderList.index("RxDLC")]]
			for subList in signalDataList:
				tmpList = [subList[signalHeaderList.index("RxCANID")],
						   subList[signalHeaderList.index("RxChannel")],
						   subList[signalHeaderList.index("RxDTC")],
						   subList[signalHeaderList.index("RxPeriod")],
						   subList[signalHeaderList.index("RxDLC")]]
				if tmpList not in self.signalValidDataList:
					self.signalValidDataList.append(tmpList)
			# 将源信号数据先以ID大小排序，再以通道大小排序
			self.signalValidDataList = sorted(self.signalValidDataList, key=lambda subList:[int(subList[self.sigValidHeaderList.index("RxCANID")], 16),\
																							int(subList[self.sigValidHeaderList.index("RxChannel")], 16)])
			# print(self.signalValidDataList)
			# print(len(self.signalValidDataList))

	# 普通报文数据处理
	def __msgDataHandling(self, index):
		retList = []
		if self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "Qoros_C6M0" or self.config.platInfo == "CHJ":
			retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("TxCANID")]) # ID
			retList.append(str(len(self.msgDesChList[index]) - 1) + 'u') # dest_mo_num
			retList.append("255u") # src_ecu_node
			retList.append("255u") # valid_flg_index(DTC)
			retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("TxDLC")] + 'u') # DLC
			if self.msgValidDataList[index][self.msgValidHeaderList.index("RxInterrupt")] == '0': # dest_mo
				for ch in sorted(self.msgDesChList[index][1:]):
					retList.append(ch + 'u')
					retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RouteCondiction")] + 'u')
				for i in range(5 - len(self.msgDesChList[index]) + 1):
					retList.append('0u')
					retList.append('0u')
			else:
				for ch in sorted(self.msgDesChList[index][1:]):
					retList.append('0u')
					retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RouteCondiction")] + 'u')
				for i in range(5 - len(self.msgDesChList[index]) + 1):
					retList.append('0u')
					retList.append('0u')
			if(self.msgValidDataList[index][self.msgValidHeaderList.index("TxPeriod")] != "None"):	
				retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("TxPeriod")] + 'u') # cycle
			else:
				retList.append('0u')
			retList.append('0u') # msg_index
			retList.append('0x0000u') # buf_index
			retList.append('0u') # pre_callback
			retList.append('0u') # post_callback
			retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RxInterrupt")]) # RxInterrupt
			retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RxDTC")]) # RxDTC
			retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RxChannel")]) # RxCh
		elif self.config.platInfo == "GAW1.2_OldPlatform":
			retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("TxCANID")]) # ID
			retList.append(str(len(self.msgDesChList[index]) - 1) + 'u') # dest_mo_num
			# retList.append("255u") # src_ecu_node
			retList.append("255u") # valid_flg_index(DTC)
			# retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("TxDLC")] + 'u') # DLC
			if self.msgValidDataList[index][self.msgValidHeaderList.index("RxInterrupt")] == '0': # dest_mo
				for ch in sorted(self.msgDesChList[index][1:]):
					retList.append(ch + 'u')
					retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RouteCondiction")] + 'u')
				for i in range(5 - len(self.msgDesChList[index]) + 1):
					retList.append('0u')
					retList.append('0u')
			else:
				for ch in sorted(self.msgDesChList[index][1:]):
					retList.append('0u')
					retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RouteCondiction")] + 'u')
				for i in range(5 - len(self.msgDesChList[index]) + 1):
					retList.append('0u')
					retList.append('0u')
			# if(self.msgValidDataList[index][self.msgValidHeaderList.index("TxPeriod")] != "None"):	
			# 	retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("TxPeriod")] + 'u') # cycle
			# else:
			# 	retList.append('0u')
			retList.append('0u') # msg_index
			retList.append('0x0000u') # buf_index
			# retList.append('0u') # pre_callback
			# retList.append('0u') # post_callback
			retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RxInterrupt")]) # RxInterrupt
			retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RxDTC")]) # RxDTC
			retList.append(self.msgValidDataList[index][self.msgValidHeaderList.index("RxChannel")]) # RxCh


		# print(retList)
		return retList

	# 信号报文数据处理
	def __sigDataHandling(self, index):
		retList = []
		if self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "Qoros_C6M0" or self.config.platInfo == "CHJ":
			retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxCANID")]) # ID
			retList.append("0u") # dest_mo_num
			retList.append("255u") # src_ecu_node
			retList.append("255u") # valid_flg_index(DTC)
			retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxDLC")] + 'u') # DLC
			for i in range(5): # dest_mo
				retList.append('0u')
				retList.append('0u')
			retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxPeriod")] + 'u') # cycle
			retList.append('0u') # msg_index
			retList.append('0x0000u') # buf_index
			retList.append("0u") # pre_callback
			retList.append("0u") # post_callback
			retList.append('0') # RxInterrupt
			retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxDTC")]) # RxDTC
			retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxChannel")]) # RxCh
		elif self.config.platInfo == "GAW1.2_OldPlatform":
			retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxCANID")]) # ID
			retList.append("0u") # dest_mo_num
			# retList.append("255u") # src_ecu_node
			retList.append("255u") # valid_flg_index(DTC)
			# retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxDLC")] + 'u') # DLC
			for i in range(5): # dest_mo
				retList.append('0u')
				retList.append('0u')
			# retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxPeriod")] + 'u') # cycle
			retList.append('0u') # msg_index
			retList.append('0x0000u') # buf_index
			# retList.append("0u") # pre_callback
			# retList.append("0u") # post_callback
			retList.append('0') # RxInterrupt
			retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxDTC")]) # RxDTC
			retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxChannel")]) # RxCh

		# print(retList)
		return retList

	def data_handle(self):
		"""数据处理"""
		# 将报文路由存入self.RouterTable中
		for index in range(len(self.msgValidDataList)):
			self.routerTableList.append(self.__msgDataHandling(index))
		# print(self.routerTableList)
		# print(len(self.routerTableList))
		# print(self.msgValidDataListISR)

		# 将信号源报文存入self.routerTableList
		for index in range(len(self.signalValidDataList)):
			# 轮询普通报文列表查询当前信号报文ID是否存在于普通报文转发列表
			# if len(list(filter(lambda subList: subList[self.msgValidHeaderList.index("TxCANID")] == self.signalValidDataList[index][self.sigValidHeaderList.index("RxCANID")], self.msgValidDataList))) == 0:
			if len(list(filter(lambda subList: [subList[self.msgValidHeaderList.index("RxCANID")], subList[self.msgValidHeaderList.index("RxChannel")]] == \
												[self.signalValidDataList[index][self.sigValidHeaderList.index("RxCANID")],self.signalValidDataList[index][self.sigValidHeaderList.index("RxChannel")]], \
												self.msgValidDataList))) == 0:
				self.routerTableList.append(self.__sigDataHandling(index))
		# print(self.routerTableList)
		# print(len(self.routerTableList))

		# 将所有报文按ID大小排序, 在ID排序基础上以接收通道排序；所有表中都这样处理。
		self.routerTableList = sorted(self.routerTableList, key=lambda subList: [int(subList[self.routerTableListHeader.index("RxCANID")], 16), int(subList[self.routerTableListHeader.index("RxChannel")])])
		# print(self.routerTableList)
		# print(len(self.routerTableList))
		
		# 对信号源报文进行处理
		SrcSig_Num = 0
		for subSigList in self.signalValidDataList:
			for subRouterList in self.routerTableList:
				# print(subRouterList)
				# if subSigList[self.sigValidHeaderList.index("RxCANID")] == subRouterList[self.routerTableListHeader.index("RxCANID")]:
				if [subSigList[self.sigValidHeaderList.index("RxCANID")], subSigList[self.sigValidHeaderList.index("RxChannel")]] == \
				[subRouterList[self.routerTableListHeader.index("RxCANID")], subRouterList[self.routerTableListHeader.index("RxChannel")]]:
					subRouterList[self.routerTableListHeader.index("msg_index")] = str(SrcSig_Num) + 'u'
					subRouterList[self.routerTableListHeader.index("buf_index")] = "0x" + hex(int("0xF000", 16) + int(subRouterList[self.routerTableListHeader.index("buf_index")][:-1], 16) + 8*SrcSig_Num)[2:].upper() + 'u'
					SrcSig_Num += 1

		# 对判别DTC报文进行标序,并且将路由表拆分为中断和FIFO
		DTC_Num = 0
		for subList in self.routerTableList:		
			if subList[self.routerTableListHeader.index("RxDTC")] == "Y":
				subList[self.routerTableListHeader.index("valid_flg_index")] = str(DTC_Num) + 'u'
				DTC_Num += 1
			
			if subList[self.routerTableListHeader.index("RxInterrupt")] == '1':
				self.routerTableISRList.append(subList)
			else:
				self.routerTableFIFOList.append(subList)
			# print(subList)
		# print(self.routerTableISRList)
		
		# 对中断路由报文发送MO进行再处理
		MO_ISR_num = len(self.routerTableISRList)
		for i in range(len(self.routerTableISRList)):
			for j in range(len(self.msgDesChListISR[i][1:])):
				self.routerTableISRList[i][self.routerTableListHeader.index("dest_mo_index1") + j*2] = str(MO_ISR_num) + 'u'
				MO_ISR_num += 1
		# print(self.routerTableISRList)

	def build_table(self):
		"""创建数组表"""
		pass


# 中断轮询表
class PbDirectRoutingTable(RoutingTable, HexBase):
	"""PB_DirectRoutingTable"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.msgValidDataList = []
		self.msgValidDataListISR = []
		self.msgDesChList = []
		self.msgDesChListISR = []
		self.signalValidDataList = []
		self.routerTableList = []
		if self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "Qoros_C6M0" or self.config.platInfo == "CHJ":
			self.routerTableListHeader = ["RxCANID", "dest_mo_num", "src_ecu_node", "valid_flg_index", "DLC", "dest_mo_index1", "dest_mo_condition1",\
										  "dest_mo_index2", "dest_mo_condition2", "dest_mo_index3", "dest_mo_condition3", "dest_mo_index4", "dest_mo_condition4",\
										  "dest_mo_index5", "dest_mo_condition5", "cycle", "msg_index", "buf_index", "pre_callback", "post_callback",\
										  "RxInterrupt", "RxDTC", "RxChannel"]
		elif self.config.platInfo == "GAW1.2_OldPlatform":
			self.routerTableListHeader = ["RxCANID", "dest_mo_num",  "valid_flg_index", "dest_mo_index1", "dest_mo_condition1",\
									  "dest_mo_index2", "dest_mo_condition2", "dest_mo_index3", "dest_mo_condition3", "dest_mo_index4", "dest_mo_condition4",\
									  "dest_mo_index5", "dest_mo_condition5", "msg_index", "buf_index",\
									  "RxInterrupt", "RxDTC", "RxChannel"]
										  
		self.routerTableISRList = []
		self.routerTableFIFOList = []
		self.PB_DirectRoutingTable = []
		if self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "Qoros_C6M0" or self.config.platInfo == "CHJ":	
			self.structType = ["uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint16", "uint16", "uint16", "uint8", "uint8"]
			self.structLen = 22
		elif self.config.platInfo == "GAW1.2_OldPlatform":
			self.structType = ["uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint16", "uint16"]
			self.structLen = 16
		# self.tableLenAddr = "0x00c08001"
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["PB_DirectRoutingTable"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["PB_DirectRoutingTable"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c094be"
		# self.tableLen = 20
		self.tableAddr = self.config.addrInfo["PB_DirectRoutingTable"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["PB_DirectRoutingTable"]["tableLen"])
		self.hexDataList = []

	# 创建PB_DirectRoutingTable数组
	def build_table(self):
		"""创建数组表"""
		self.PB_DirectRoutingTable.append("/*------------------table2: PB_DirectRoutingTable -----------------------------*/\n")
		self.PB_DirectRoutingTable.append("const uint8 PB_Table0_Len __at(0x00c08001) = " + str(len(self.routerTableISRList)) + "u;\n")
		self.PB_DirectRoutingTable.append("const PB_TABLE_1 PB_DirectRoutingTable[TABLE_0_SIZE] __at(0x00c094be)\n")
		self.PB_DirectRoutingTable.append("=\n")
		self.PB_DirectRoutingTable.append("{\n")
		for index, subList in enumerate(self.routerTableISRList):
			# print(subList)
			self.PB_DirectRoutingTable.append('/*' + subList[0] + '*/' + '{')
			self.PB_DirectRoutingTable.append(','.join(subList[self.routerTableListHeader.index("dest_mo_num"):self.routerTableListHeader.index("RxInterrupt")]))
			self.PB_DirectRoutingTable.append('},' + "/*" + str(index) + "*/" + '\n')
		else:
			# 如果列表为空，则以0，填充数组，以防工程编译时不通过.
			if not self.routerTableISRList:
				self.PB_DirectRoutingTable.append('\t0,\n')
		self.PB_DirectRoutingTable.append("};")


# 报文轮询表
class PbMsgRoutingTable(RoutingTable, HexBase):
	"""PB_MsgRoutingTable"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.msgValidDataList = []
		self.msgValidDataListISR = []
		self.msgDesChList = []
		self.msgDesChListISR = []
		self.signalValidDataList = []
		self.routerTableList = []
		if self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "Qoros_C6M0" or self.config.platInfo == "CHJ":
			self.routerTableListHeader = ["RxCANID", "dest_mo_num", "src_ecu_node", "valid_flg_index", "DLC", "dest_mo_index1", "dest_mo_condition1",\
										  "dest_mo_index2", "dest_mo_condition2", "dest_mo_index3", "dest_mo_condition3", "dest_mo_index4", "dest_mo_condition4",\
										  "dest_mo_index5", "dest_mo_condition5", "cycle", "msg_index", "buf_index", "pre_callback", "post_callback",\
										  "RxInterrupt", "RxDTC", "RxChannel"]
		elif self.config.platInfo == "GAW1.2_OldPlatform":
			self.routerTableListHeader = ["RxCANID", "dest_mo_num",  "valid_flg_index", "dest_mo_index1", "dest_mo_condition1",\
									  "dest_mo_index2", "dest_mo_condition2", "dest_mo_index3", "dest_mo_condition3", "dest_mo_index4", "dest_mo_condition4",\
									  "dest_mo_index5", "dest_mo_condition5", "msg_index", "buf_index",\
									  "RxInterrupt", "RxDTC", "RxChannel"]


		self.routerTableISRList = []
		self.routerTableFIFOList = []
		self.PB_MsgRoutingTable = []
		if self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "Qoros_C6M0" or self.config.platInfo == "CHJ":	
			self.structType = ["uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint16", "uint16", "uint16", "uint8", "uint8"]
			self.structLen = 22
		elif self.config.platInfo == "GAW1.2_OldPlatform":
			self.structType = ["uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint16", "uint16"]
			self.structLen = 16
		# self.tableLenAddr = "0x00c0a46d"
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["PB_MsgRoutingTable"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["PB_MsgRoutingTable"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c14200"
		# self.tableLen = 256
		self.tableAddr = self.config.addrInfo["PB_MsgRoutingTable"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["PB_MsgRoutingTable"]["tableLen"])
		self.hexDataList = []

		

		


	# 创建PB_MsgRoutingTable数组
	def build_table(self):
		"""创建数组表"""
		self.PB_MsgRoutingTable.append("/*------------------table3: PB_MsgRoutingTable -----------------------------*/\n")
		self.PB_MsgRoutingTable.append("const uint8 PB_Table1_Len __at(0x00c0a46d) = " + str(len(self.routerTableFIFOList)) + "u;\n")
		self.PB_MsgRoutingTable.append("const PB_TABLE_1 SHUGE PB_MsgRoutingTable[TABLE_1_SIZE] __at(0x00c14200)\n")
		self.PB_MsgRoutingTable.append("=\n")
		self.PB_MsgRoutingTable.append("{\n")
		for index, subList in enumerate(self.routerTableFIFOList):
			# print(subList)
			self.PB_MsgRoutingTable.append('/*' + subList[self.routerTableListHeader.index("RxCANID")] + ": NODE_" + get_column_letter(int(subList[self.routerTableListHeader.index("RxChannel")])) + '*/' + '{')
			self.PB_MsgRoutingTable.append(','.join(subList[self.routerTableListHeader.index("dest_mo_num"):self.routerTableListHeader.index("RxInterrupt")]))
			self.PB_MsgRoutingTable.append('},' + "/*" + str(index) + "*/" + '\n')
		else:
			# 如果列表为空，则以0，填充数组，以防工程编译时不通过.
			if not self.routerTableFIFOList:
				self.PB_MsgRoutingTable.append('\t0,\n')
		self.PB_MsgRoutingTable.append("};")


# 信号报文接收表
class PbMsgRecvTable(HexBase):
	"""PB_Msg_Recv_Table"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.signalValidDataList = []
		self.PBMsgRecvTableList = []
		self.PB_Msg_Recv_Table = []
		
		self.structType = ["uint32", "uint16", "uint16", "uint8", "uint8", "uint8", "uint8"]
		self.structLen = 12
		# self.tableLenAddr = "0x00c0a51f"
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["PB_Msg_Recv_Table"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["PB_Msg_Recv_Table"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c07cf0"
		# self.tableLen = 64
		self.tableAddr = self.config.addrInfo["PB_Msg_Recv_Table"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["PB_Msg_Recv_Table"]["tableLen"])
		self.hexDataList = []

		

		

	def get_valid_data(self, signalRoute):
		"""获取有效数据"""
		# 深拷贝一份路由数据，不破坏原有数据
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)
		
		if signalDataList:
		# 提取信号有效头列表		
			self.sigValidHeaderList = [signalHeaderList[signalHeaderList.index("RxCANID")],\
									   signalHeaderList[signalHeaderList.index("RxPeriod")],\
									   signalHeaderList[signalHeaderList.index("RxDLC")],\
									   signalHeaderList[signalHeaderList.index("RxChannel")],\
									   signalHeaderList[signalHeaderList.index("ByteOrder")]]

			for subList in signalDataList:
				tmpList = [subList[signalHeaderList.index("RxCANID")],
						   subList[signalHeaderList.index("RxPeriod")],
						   subList[signalHeaderList.index("RxDLC")],
						   subList[signalHeaderList.index("RxChannel")],
						   subList[signalHeaderList.index("ByteOrder")]]
				if tmpList not in self.signalValidDataList:
					self.signalValidDataList.append(tmpList)
			self.signalValidDataList = sorted(self.signalValidDataList, key=lambda subList:[int(subList[self.sigValidHeaderList.index("RxCANID")], 16),\
																							int(subList[self.sigValidHeaderList.index("RxChannel")], 16)])
			# print(self.signalValidDataList)
			# print(len(self.signalValidDataList))

	def __src_signal_handling(self, index):
		"""对源信号报文进行处理，得到PB_Msg_Recv_Table表的一行"""
		retList = []
		retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxCANID")] + 'u')
		retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxPeriod")] + 'u')
		retList.append(str(int(self.signalValidDataList[index][self.sigValidHeaderList.index("RxPeriod")]) * 5) + 'u')
		retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxChannel")] + 'u')
		retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("RxDLC")] + 'u')
		retList.append(self.signalValidDataList[index][self.sigValidHeaderList.index("ByteOrder")] + 'u')
		retList.append('0u')

		# print(retList)
		return retList


	def data_handle(self):
		"""数据处理"""
		for index in range(len(self.signalValidDataList)):
			self.PBMsgRecvTableList.append(self.__src_signal_handling(index))

	def build_table(self):
		"""创建数组表"""
		self.PB_Msg_Recv_Table.append("/*------------------table4: PB_Msg_Recv_Table -----------------------------*/\n")
		self.PB_Msg_Recv_Table.append("const uint8 PB_Table2_Len __at(0x00c0a51f) = " + str(len(self.PBMsgRecvTableList)) + "u;\n")
		self.PB_Msg_Recv_Table.append("const PB_TAB_RX_TYP SHUGE PB_Msg_Recv_Table[TABLE_2_SIZE] __at(0x00c07cf0)\n")
		self.PB_Msg_Recv_Table.append("=\n")
		self.PB_Msg_Recv_Table.append("{\n")
		for index, subList in enumerate(self.PBMsgRecvTableList):
			self.PB_Msg_Recv_Table.append("/*" + "NODE_" + get_column_letter(int(self.signalValidDataList[index][self.sigValidHeaderList.index("RxChannel")])) + "*/")
			self.PB_Msg_Recv_Table.append("{")
			self.PB_Msg_Recv_Table.append(','.join(subList))
			self.PB_Msg_Recv_Table.append('},' + "/*" + str(index) + "*/" + '\n')
		else:
			# 如果列表为空，则以0，填充数组，以防工程编译时不通过.
			if not self.PBMsgRecvTableList:
				self.PB_Msg_Recv_Table.append('\t0,\n')
		self.PB_Msg_Recv_Table.append("};")


# 信号路由表
class PbSignalRoutingTable(HexBase):
	"""PB_Signal_Routing_Table"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.txCanIdList = []
		self.rxCanIdList = []
		self.validDataListList = []
		self.srcSignalListList = []
		self.PbSignalRoutingTableList = []
		self.PB_Signal_Routing_Table = []
		if self.config.platInfo == "GAW1.2_OldPlatform" or self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "CHJ":	
			self.structType = ["uint16", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8"]
			self.structLen = 8
		elif self.config.platInfo == "Qoros_C6M0":	
			self.structType = ["uint16", "uint8", "uint8", "uint16", "uint8", "uint8", "uint8", "uint8"]
			self.structLen = 10
		# self.tableLenAddr = ""
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["PB_Signal_Routing_Table"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["PB_Signal_Routing_Table"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c12200"
		# self.tableLen = 1024
		self.tableAddr = self.config.addrInfo["PB_Signal_Routing_Table"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["PB_Signal_Routing_Table"]["tableLen"])
		self.hexDataList = []

	def get_valid_data(self, signalRoute):
		"""获取有效数据"""
		# 深拷贝一份路由数据，不破坏原有数据
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)

		if signalDataList:
			# 提取信号有效头列表
			self.sigValidHeaderList = signalHeaderList

			# 首先获取发送信号ID，并以ID大小进行排序, 另区分不同通道，在ID排序的基础上以通道排序
			for subList in signalDataList:
				txCanIdCh = [subList[self.sigValidHeaderList.index("TxCANID")], subList[self.sigValidHeaderList.index("TxChannle")], subList[self.sigValidHeaderList.index("TxPeriod")]]
				if txCanIdCh not in self.txCanIdList:
					self.txCanIdList.append(txCanIdCh)
			self.txCanIdList = sorted(self.txCanIdList, key=lambda txCanIdCh:[int(txCanIdCh[2]), int(txCanIdCh[0], 16), int(txCanIdCh[1])])
			# print(self.txCanIdList)

			# 然后根据已排序的发送信号ID获取有效信号数据，存储在列表中(相当于对整个信号表数据按发送ID大小进行按块排序)
			# 同时获取每个目标发送ID对应的源信号列表，用于将信号路由表按目标信号分类
			for txCanIdCh in self.txCanIdList:
				# 每一个发送ID对应的接收ID存储在一个列表中，最后再将所有rxCanIdList存储在一个列表中
				# validDataList = []
				srcSignalList = []
				for subList in signalDataList:
					if [subList[self.sigValidHeaderList.index("TxCANID")], subList[self.sigValidHeaderList.index("TxChannle")], \
						subList[self.sigValidHeaderList.index("TxPeriod")]] == txCanIdCh:
						# validDataList.append(subList)
						# print(subList)
						self.validDataListList.append(subList)
						srcSignalList.append(subList)
				self.srcSignalListList.append(srcSignalList)	

			# 获取接收信号ID，并以ID大小排序，用于索引接收信号表(PbMsgRecvTable)的字节位位置
			for subList in signalDataList:
				rxCanIdCh = [subList[self.sigValidHeaderList.index("RxCANID")], subList[self.sigValidHeaderList.index("RxChannel")]]
				if rxCanIdCh not in self.rxCanIdList:
					self.rxCanIdList.append(rxCanIdCh)
			self.rxCanIdList = sorted(self.rxCanIdList, key=lambda rxCanIdCh:[int(rxCanIdCh[0], 16), int(rxCanIdCh[1])])
			# print(self.rxCanIdList)

	def __calculate_source_index(self, subList) -> str:
		"""计算源信号字节索引"""
		ret = str(self.rxCanIdList.index([subList[self.sigValidHeaderList.index("RxCANID")], subList[self.sigValidHeaderList.index("RxChannel")]]) * 8 + \
			int(subList[self.sigValidHeaderList.index("RxStartBit")])//8)

		return ret


	def __signal_route_handling(self, subList):
		"""对获取的每一行数据处理，得到信号路由表(PB_Signal_Routing_Table)的一行"""
		retList = []
		if self.config.platInfo == "GAW1.2_OldPlatform" or self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "CHJ":
			retList.append(self.__calculate_source_index(subList) + 'u')
			retList.append(str(int(subList[self.sigValidHeaderList.index("RxStartBit")])%8) + 'u')
			retList.append(subList[self.sigValidHeaderList.index("RxSigLen")] + 'u')
			retList.append(str(int(subList[self.sigValidHeaderList.index("TxStartBit")])//8) + 'u')
			retList.append(str(int(subList[self.sigValidHeaderList.index("TxStartBit")])%8) + 'u')
			retList.append(subList[self.sigValidHeaderList.index("ByteOrder")] + 'u')
			retList.append('0xffu')
		elif self.config.platInfo == "Qoros_C6M0":	
			retList.append(self.__calculate_source_index(subList) + 'u')
			retList.append(str(int(subList[self.sigValidHeaderList.index("RxStartBit")])%8) + 'u')
			retList.append(subList[self.sigValidHeaderList.index("RxSigLen")] + 'u')
			retList.append(subList[self.sigValidHeaderList.index("ByteOrder")] + 'u')
			retList.append(str(int(subList[self.sigValidHeaderList.index("TxStartBit")])//8) + 'u')
			retList.append(str(int(subList[self.sigValidHeaderList.index("TxStartBit")])%8) + 'u')	
			retList.append(subList[self.sigValidHeaderList.index("TxByteOrder")] + 'u')
			retList.append('0xffu')


		# print(retList)
		return retList

	def data_handle(self):
		"""数据处理"""
		for subList in self.validDataListList:
			# for subList in validDataList:
			self.PbSignalRoutingTableList.append(self.__signal_route_handling(subList))

	def build_table(self):
		"""创建数组表"""
		self.PB_Signal_Routing_Table.append("/*------------------table5: PB_Signal_Routing_Table -----------------------------*/\n")
		self.PB_Signal_Routing_Table.append("const  PB_TABLE_3 SHUGE PB_Signal_Routing_Table[TABLE_3_SIZE] __at(0x00c12200)\n")
		self.PB_Signal_Routing_Table.append("=\n")
		self.PB_Signal_Routing_Table.append("{\n")
		self.PB_Signal_Routing_Table.append("\t/******** Signal Number: " + str(len(self.PbSignalRoutingTableList)) + " ********/" + "\n")
		for index in range(len(self.PbSignalRoutingTableList)):
			# 以每一个发送ID为一个块，将信号路由表分开，方便查看信息
			# 思路：根据每一个发送ID对应的接收信号数量，与全部信号索引比较，当信号索引等于发送信号对应源信号数量和时，
			#       打印一行发送信号的信息 CANID、周期、发送通道，对应源信号数量.			
			num = 0
			start_flag = False
			for i in range(len(self.srcSignalListList)):
				if index == num:
					start_flag = True
					break
				num += len(self.srcSignalListList[i])
			if start_flag == True:
				self.PB_Signal_Routing_Table.append("/******** " + self.validDataListList[index][self.sigValidHeaderList.index("TxCANID")] + " "*5 +\
																	self.validDataListList[index][self.sigValidHeaderList.index("TxPeriod")] + " "*5 + \
																	"NODE_" + get_column_letter(int(self.validDataListList[index][self.sigValidHeaderList.index("TxChannle")])) + " "*5 + \
																	str(len(self.srcSignalListList[i])) + " ********/" + "\n")

			self.PB_Signal_Routing_Table.append("/*" + self.validDataListList[index][self.sigValidHeaderList.index("RxCANID")] + "*/")
			self.PB_Signal_Routing_Table.append("{")
			self.PB_Signal_Routing_Table.append(','.join(self.PbSignalRoutingTableList[index]))
			self.PB_Signal_Routing_Table.append("},")
			self.PB_Signal_Routing_Table.append("/* " + "NODE_" + get_column_letter(int(self.validDataListList[index][self.sigValidHeaderList.index("RxChannel")])) + \
												"   " + self.validDataListList[index][self.sigValidHeaderList.index("SignalName")] + " */")
			self.PB_Signal_Routing_Table.append("/*" + str(index) + "*/" + '\n')
		else:
			# 如果列表为空，则以0，填充数组，以防工程编译时不通过.
			if not self.PbSignalRoutingTableList:
				self.PB_Signal_Routing_Table.append('\t0,\n')
		self.PB_Signal_Routing_Table.append("};")
		

# 信号报文发送表
class PbMsgSendTable(HexBase):
	"""PB_Msg_Send_Table"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.txCanIdList = []
		self.validDataList = []
		self.srcSignalListList = [] # 同一ID发送信号对应的源信号数量
		# self.rxCanIdList = []
		self.txCanIdrxCanIdCountList = []
		self.PbMsgSendTableList = []
		self.PB_Msg_Send_Table = []	
		if self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "Qoros_C6M0":
			self.structType = ["uint32", "uint16", "uint8", "uint8", "uint8", "uint8", "uint16", "uint8", "uint8"] 
			self.structLen = 14
		elif self.config.platInfo == "GAW1.2_OldPlatform":
			self.structType = ["uint32", "uint16", "uint8", "uint8", "uint8", "uint8", "uint16"]
			self.structLen = 12
		elif self.config.platInfo == "CHJ":
			self.structType = ["uint32", "uint16", "uint8", "uint8", "uint8", "uint8", "uint16","uint16"] 
			self.structLen = 14
		# self.tableLenAddr = ""
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["PB_Msg_Send_Table"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["PB_Msg_Send_Table"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c08002"
		# self.tableLen = 128
		self.tableAddr = self.config.addrInfo["PB_Msg_Send_Table"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["PB_Msg_Send_Table"]["tableLen"])
		self.hexDataList = []

		

		

	def get_valid_data(self, signalRoute):
		"""获取有效数据"""
		# 深拷贝一份路由数据，不破坏原有数据
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)

		if signalDataList:
			# 获取发送信号ID，并以ID大小进行排序, 加通道约束，区分不同通道
			for subList in signalDataList:
				txCanIdXCh = [subList[signalHeaderList.index("TxCANID")], subList[signalHeaderList.index("TxChannle")], subList[signalHeaderList.index("TxPeriod")]]
				if txCanIdXCh not in self.txCanIdList:
					self.txCanIdList.append(txCanIdXCh)
			self.txCanIdList = sorted(self.txCanIdList, key=lambda txCanIdXCh:[int(txCanIdXCh[2]), int(txCanIdXCh[0], 16), int(txCanIdXCh[1])])
			# print(self.txCanIdList)

			# 提取信号有效头列表
			self.sigValidHeaderList = [signalHeaderList[signalHeaderList.index("TxCANID")],\
									   signalHeaderList[signalHeaderList.index("TxPeriod")],\
									   signalHeaderList[signalHeaderList.index("TxDLC")],\
									   signalHeaderList[signalHeaderList.index("TxChannle")],\
									   signalHeaderList[signalHeaderList.index("ByteOrder")]]

			# 获取发送信号数据，先以周期进行排序的基础上以ID大小进行排序,最后以CH大小排序
			for subList in signalDataList:
				txList = [subList[signalHeaderList.index("TxCANID")],
						  subList[signalHeaderList.index("TxPeriod")],
						  subList[signalHeaderList.index("TxDLC")],
						  subList[signalHeaderList.index("TxChannle")],
						  subList[signalHeaderList.index("ByteOrder")]]
				if txList not in self.validDataList:
					self.validDataList.append(txList)
			self.validDataList = sorted(self.validDataList, key=lambda subList:[int(subList[self.sigValidHeaderList.index("TxPeriod")]), \
										int(subList[self.sigValidHeaderList.index("TxCANID")], 16), int(subList[self.sigValidHeaderList.index("TxChannle")])])
			# print(self.validDataList)

			# 然后根据已排序的发送信号ID获取对应的源信号，主要想获得对应数量
			for subvalidDataList in self.validDataList:
				# 每一个发送ID对应的接收ID存储在一个列表中，最后再将对应所有接收数据存储在一个列表中
				srcSignalList = []
				for subList in signalDataList:
					if [subList[signalHeaderList.index("TxCANID")], subList[signalHeaderList.index("TxChannle")]] \
					== [subvalidDataList[self.sigValidHeaderList.index("TxCANID")], subvalidDataList[self.sigValidHeaderList.index("TxChannle")]]:
						srcSignalList.append(subList)
				# print(len(srcSignalList))
				self.srcSignalListList.append(srcSignalList)

			# 提取信号发送报文ID和对应源信号数量，先以发送周期排序，再以ID大小进行排序，最后以CH大小排序
			for index in range(len(self.validDataList)):
				tmpList = [self.validDataList[index][self.sigValidHeaderList.index("TxCANID")], self.validDataList[index][self.sigValidHeaderList.index("TxChannle")], \
							self.validDataList[index][self.sigValidHeaderList.index("TxPeriod")], str(len(self.srcSignalListList[index]))]
				self.txCanIdrxCanIdCountList.append(tmpList)
			self.txCanIdrxCanIdCountList = sorted(self.txCanIdrxCanIdCountList, key=lambda subList:[int(subList[2]), int(subList[0], 16), int(subList[1])])
			# print(self.txCanIdrxCanIdCountList)

			# # 获取接收信号ID，并以ID大小排序，用于索引接收信号表(PbMsgRecvTable)的字节位位置
			# for subList in signalDataList:
			# 	rxCanId = subList[column_index_from_string('J') - 1]
			# 	if rxCanId not in self.rxCanIdList:
			# 		self.rxCanIdList.append(rxCanId)
			# self.rxCanIdList = sorted(self.rxCanIdList, key=lambda rxCanId:int(rxCanId, 16))
			# print(self.rxCanIdList)

	# 计算发送信号对应的第一个源信号索引
	def __calculate_first_source_index(self, index) -> str:
		ret = 0
		for i in range(self.txCanIdList.index([self.validDataList[index][self.sigValidHeaderList.index("TxCANID")], \
												self.validDataList[index][self.sigValidHeaderList.index("TxChannle")], self.validDataList[index][self.sigValidHeaderList.index("TxPeriod")]])):
			ret += int(self.txCanIdrxCanIdCountList[i][3])

		return str(ret)


	def __send_signal_data_handling(self, index):
		"""对发送信号ID进行处理，提取信号发送表的子元素"""
		retList = []
		# print(self.config.platInfo)
		if self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "Qoros_C6M0":
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxCANID")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxPeriod")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxChannle")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxDLC")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("ByteOrder")] + 'u')
			retList.append(str(len(self.srcSignalListList[index])) + 'u')
			retList.append(self.__calculate_first_source_index(index) + 'u')
			retList.append('0u')
			retList.append('0u')
		elif self.config.platInfo == "GAW1.2_OldPlatform":
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxCANID")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxPeriod")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxChannle")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxDLC")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("ByteOrder")] + 'u')
			retList.append(str(len(self.srcSignalListList[index])) + 'u')
			retList.append(self.__calculate_first_source_index(index) + 'u')
			# retList.append('0u')
			# retList.append('0u')
		elif self.config.platInfo == "CHJ":
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxCANID")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxPeriod")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxChannle")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("TxDLC")] + 'u')
			retList.append(self.validDataList[index][self.sigValidHeaderList.index("ByteOrder")] + 'u')
			retList.append(str(len(self.srcSignalListList[index])) + 'u')
			retList.append(self.__calculate_first_source_index(index) + 'u')
			retList.append('1u')  # checksum
			# retList.append('0u')


		# print(retList)
		return retList		


	def data_handle(self):
		"""数据处理"""
		for index in range(len(self.validDataList)):
			self.PbMsgSendTableList.append(self.__send_signal_data_handling(index))


	def build_table(self):
		"""创建数组表"""
		self.PB_Msg_Send_Table.append("/*------------------table6: PB_Msg_Send_Table -----------------------------*/\n")
		self.PB_Msg_Send_Table.append("const PB_TAB_TX_TYP PB_Msg_Send_Table[TABLE_4_SIZE] __at(0x00c08002) =\n")
		self.PB_Msg_Send_Table.append("{\n")
		for index, subList in enumerate(self.PbMsgSendTableList):
			self.PB_Msg_Send_Table.append("/*" + "NODE_" + get_column_letter(int(self.validDataList[index][self.sigValidHeaderList.index("TxChannle")])) + "*/")
			self.PB_Msg_Send_Table.append("{")
			self.PB_Msg_Send_Table.append(','.join(subList))
			self.PB_Msg_Send_Table.append('},' + "/*" + str(index) + "*/" + '\n')
		else:
			# 如果列表为空，则以0，填充数组，以防工程编译时不通过.
			if not self.PbMsgSendTableList:
				self.PB_Msg_Send_Table.append('\t0,\n')
		self.PB_Msg_Send_Table.append("};")


# 目标信号对应源信号ID索引
class PbMsgSrcTable(HexBase):
	"""PB_Msg_Src_Table"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.txCanIdPeriodList = []
		self.srcSignalListList = [] # 同一ID发送信号对应的源信号表索引
		self.recvSignalList = []
		self.PbMsgSrcTableList = []
		self.PB_Msg_Src_Table = []
		self.structType = ["uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8"]
		self.structLen = 16
		# self.tableLenAddr = ""
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["PB_Msg_Src_Table"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["PB_Msg_Src_Table"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c17200"
		# self.tableLen = 128
		self.tableAddr = self.config.addrInfo["PB_Msg_Src_Table"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["PB_Msg_Src_Table"]["tableLen"])
		self.hexDataList = []

	def get_valid_data(self, signalRoute):
		"""获取有效数据"""
		# 深拷贝一份路由数据，不破坏原有数据
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)
		if signalDataList:
			# 提取信号有效头列表
			self.sigValidHeaderList = [signalHeaderList[signalHeaderList.index("TxCANID")],\
									   signalHeaderList[signalHeaderList.index("TxPeriod")],\
									   signalHeaderList[signalHeaderList.index("TxChannle")]]

			# 获取发送信号数据，先以周期进行排序的基础上以ID大小进行排序
			for subList in signalDataList:
				txList = [subList[signalHeaderList.index("TxCANID")],\
						  subList[signalHeaderList.index("TxPeriod")],\
						  subList[signalHeaderList.index("TxChannle")]]
				if txList not in self.txCanIdPeriodList:
					self.txCanIdPeriodList.append(txList)
			self.txCanIdPeriodList = sorted(self.txCanIdPeriodList, key=lambda subList:[int(subList[self.sigValidHeaderList.index("TxPeriod")]), \
																						int(subList[self.sigValidHeaderList.index("TxCANID")], 16),\
																						int(subList[self.sigValidHeaderList.index("TxChannle")])])
			# print(self.txCanIdPeriodList)
			# print(len(self.txCanIdPeriodList))

			# 然后根据已排序的发送信号ID获取对应的源信号，主要想获得对应于源信号表中的索引
			for txCanId in self.txCanIdPeriodList:
				# 每一个发送ID对应的接收ID存储在一个列表中，最后再将对应所有接收数据存储在一个列表中
				srcSignalList = []
				for subList in signalDataList:
					if [subList[signalHeaderList.index("TxCANID")], subList[signalHeaderList.index("TxChannle")]] == \
						[txCanId[self.sigValidHeaderList.index("TxCANID")], txCanId[self.sigValidHeaderList.index("TxChannle")]]:
						if [subList[signalHeaderList.index("RxCANID")], subList[signalHeaderList.index("RxChannel")]] not in srcSignalList:
							srcSignalList.append([subList[signalHeaderList.index("RxCANID")], subList[signalHeaderList.index("RxChannel")]])
				# srcSignalList = sorted(list(set(srcSignalList)), key=lambda srcSignal:[int(srcSignal[0], 16), int(srcSignal[1])])
				srcSignalList = sorted(srcSignalList, key=lambda srcSignal:[int(srcSignal[0], 16), int(srcSignal[1])])
				# print(srcSignalList)
				self.srcSignalListList.append(srcSignalList)

			# 获取所有源信号并以ID大小进行排序
			for subList in signalDataList:
				tmpSrcId = [subList[signalHeaderList.index("RxCANID")], subList[signalHeaderList.index("RxChannel")]]
				if tmpSrcId not in self.recvSignalList:
					self.recvSignalList.append(tmpSrcId)
			self.recvSignalList = sorted(self.recvSignalList, key=lambda tmpSrcId:[int(tmpSrcId[0], 16), int(tmpSrcId[1])])

	def __src_signal_index_data_handling(self, index):
		"""对发送信号对应的源信号在源信号中的索引进行处理"""
		# retList = []
		# if self.srcSignalListList:		
		# 	# 获取发送信号对应的源信号索引放入列表中
		# 	for i in range(len(self.srcSignalListList[index])):
		# 		retList.append(str(self.recvSignalList.index(self.srcSignalListList[index][i])))
		# 	# 以发送信号对应的最多源信号量为长度对列表进行填充
		# 	for i in range(len(max(self.srcSignalListList, key=lambda subList:len(subList))) - len(self.srcSignalListList[index])):
		# 		retList.append("0xFFu")
		# else:
		# 	for i in range(16):
		# 		retList.append("0xFFu")

		retList = ["0xFFu" for i in range(16)]
		if self.srcSignalListList:		
			# 获取发送信号对应的源信号索引放入列表中
			for i in range(len(self.srcSignalListList[index])):
				retList[i] = str(self.recvSignalList.index(self.srcSignalListList[index][i])) + 'u'
				# retList.append(str(self.recvSignalList.index(self.srcSignalListList[index][i])))
			# 以发送信号对应的最多源信号量为长度对列表进行填充
			# for i in range(len(max(self.srcSignalListList, key=lambda subList:len(subList))) - len(self.srcSignalListList[index])):
			# 	retList.append("0xFFu")

		# print(retList)
		return retList	

	def data_handle(self):
		"""数据处理"""
		if self.srcSignalListList:
			for index in range(len(self.srcSignalListList)):
				self.PbMsgSrcTableList.append(self.__src_signal_index_data_handling(index))
		# else:
		# 	for index in range(0):
		# 		self.PbMsgSrcTableList.append(self.__src_signal_index_data_handling(index))

	def build_table(self):
		"""创建数组表"""
		# if self.srcSignalListList:
		# 	self.PB_Msg_Src_Table.append("const uint8 PB_Msg_Src_Table[" + str(len(self.srcSignalListList)) + "][" + \
		# 									str(len(max(self.srcSignalListList, key=lambda subList:len(subList)))) + "] =\n")
		# else:
		# 	self.PB_Msg_Src_Table.append("const uint8 PB_Msg_Src_Table[" + '16' + "][" + '16' + "] =\n")

		self.PB_Msg_Src_Table.append("const uint8 PB_Msg_Src_Table[" + 'TABLE_45_SIZE' + "][" + '16' + "] __at(0x00c17200) =\n")

		self.PB_Msg_Src_Table.append("{\n")
		for index, subList in enumerate(self.PbMsgSrcTableList):
			if self.srcSignalListList:
				if index < len(self.srcSignalListList):
					self.PB_Msg_Src_Table.append("/*" + self.txCanIdPeriodList[index][self.sigValidHeaderList.index("TxCANID")] + "*/ ")
				else:
					self.PB_Msg_Src_Table.append("/*" + "Reserved" + "*/ ")
			self.PB_Msg_Src_Table.append(','.join(subList))
			self.PB_Msg_Src_Table.append(",/*" + str(index) + "*/" + '\n')
		else:
			# 如果列表为空，则以0，填充数组，以防工程编译时不通过.
			if not self.srcSignalListList:
				self.PB_Msg_Src_Table.append('\t0,\n')
		self.PB_Msg_Src_Table.append("};")


# 信号报文发送调度表
class PbMsgSendSchedule(HexBase):
	"""PB_Msg_Send_Schedule"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.txCanIdPeriodList = []
		# self.srcSignalListList = [] # 同一ID发送信号对应的源信号表索引
		# self.recvSignalList = []
		self.PbMsgSendSchedule = []
		self.PB_Msg_Send_Schedule = []		
		self.structType = ["uint16", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8",\
									"uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8"] #第一个为uint8，字节补全处理为uint16
		self.structLen = 22
		# self.tableLenAddr = ""
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["PB_Msg_Send_Schedule"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["PB_Msg_Send_Schedule"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c09b0e"
		# self.tableLen = 10
		self.tableAddr = self.config.addrInfo["PB_Msg_Send_Schedule"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["PB_Msg_Send_Schedule"]["tableLen"])
		self.hexDataList = []

		

		

	def get_valid_data(self, signalRoute):
		"""获取有效数据"""
		# 深拷贝一份路由数据，不破坏原有数据
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)

		if signalDataList:
			# 提取信号有效头列表
			self.sigValidHeaderList = [signalHeaderList[signalHeaderList.index("TxCANID")],\
									   signalHeaderList[signalHeaderList.index("TxPeriod")],\
									   signalHeaderList[signalHeaderList.index("TxChannle")]]

			# 获取发送信号数据，先以周期进行排序的基础上以ID大小进行排序
			for subList in signalDataList:
				txList = [subList[signalHeaderList.index("TxCANID")],
						  subList[signalHeaderList.index("TxPeriod")],
						  subList[signalHeaderList.index("TxChannle")]]
				if txList not in self.txCanIdPeriodList:
					self.txCanIdPeriodList.append(txList)
			self.txCanIdPeriodList = sorted(self.txCanIdPeriodList, key=lambda subList:[int(subList[1]), int(subList[0], 16), int(subList[2])])
			# print(self.txCanIdPeriodList)
			# print(len(self.txCanIdPeriodList))

	def data_handle(self):
		"""数据处理"""
		for i in range(10):
			self.PbMsgSendSchedule.append(["0u","0xFFu","0u","0xFFu","0u","0xFFu","0u","0xFFu","0u","0xFFu","0u",\
												"0xFFu","0u","0xFFu","0u","0xFFu","0u","0xFFu","0u","0xFFu","0u"])

			self.PbMsgSendSchedule[i][0] = str(((len(self.txCanIdPeriodList) + 9 - i) // 10)) + 'u'

		
		# 根据发送信号更改调度表中的其他列
		for index in range(len(self.txCanIdPeriodList)):
			# 填充16进制为0x**
			if len(hex(index)) == 3:
				self.PbMsgSendSchedule[index%10][(index//10)*2 + 1] = hex(index)[0:2] + '0' + hex(index)[2] + 'u'
			else:
				self.PbMsgSendSchedule[index%10][(index//10)*2 + 1] = hex(index) + 'u'
			self.PbMsgSendSchedule[index%10][(index//10)*2 + 2] = str(int(self.txCanIdPeriodList[index][self.sigValidHeaderList.index("TxPeriod")])//10) + 'u'

		# print(self.PbMsgSendSchedule)


	def build_table(self):
		"""创建数组表"""
		self.PB_Msg_Send_Schedule.append("/*------------------table7: PB_Msg_Send_Schedule -----------------------------*/\n")
		self.PB_Msg_Send_Schedule.append("const PB_TABLE_SCH_TYPE PB_Msg_Send_Schedule[10] __at(0x00c09b0e) =\n")
		self.PB_Msg_Send_Schedule.append("{\n")
		for subList in self.PbMsgSendSchedule:
			self.PB_Msg_Send_Schedule.append("\t{")
			self.PB_Msg_Send_Schedule.append(subList[0] + ',')
			self.PB_Msg_Send_Schedule.append("{")
			for i in range(10):
				self.PB_Msg_Send_Schedule.append("{")
				self.PB_Msg_Send_Schedule.append(subList[i*2 + 1] + ',')
				self.PB_Msg_Send_Schedule.append(subList[i*2 + 2])
				if i != 9:
					self.PB_Msg_Send_Schedule.append("},")
				elif i == 9:
					self.PB_Msg_Send_Schedule.append("}")
			self.PB_Msg_Send_Schedule.append("}")
			self.PB_Msg_Send_Schedule.append("},\n")
		self.PB_Msg_Send_Schedule.append("};")


# 信号初始值默认值基类
class PbMsgRevInitDefaultValBase(object):
	"""Table_MsgRevInitVal && Table_MsgRevDftVal"""
	def get_valid_data(self, signalRoute):
		"""获取有效数据"""
		# 深拷贝一份路由数据，不破坏原有数据
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)
		if signalDataList:
			# 提取信号有效头列表
			self.sigValidHeaderList = signalHeaderList

			for subList in signalDataList:
				tmpList = [subList[signalHeaderList.index("RxCANID")],
						   subList[signalHeaderList.index("RxChannel")]]
				if tmpList not in self.srcSignalInfoList:
					self.srcSignalInfoList.append(tmpList)
				self.validDataList.append(subList)
			self.srcSignalInfoList = sorted(self.srcSignalInfoList, key=lambda subList:[int(subList[0], 16), int(subList[1])])
			# print(self.srcSignalInfoList)
			# print(len(self.srcSignalInfoList))

	def calc_signa_val(self, signalVal, startByte, startBit, signalLen, byteorder) -> list:
		"""计算信号值"""
		signalData = [0, 0, 0, 0, 0, 0, 0, 0]
		# 轮询信号长度的每一bit，与对应字节按位或操作
		for i in range(int(signalLen)):
			if int(byteorder) == 0:
				signalData[int(startByte) - (int(startBit) + i)//8] |= (((int(signalVal, 16) >> i) & 0x01) << ((int(startBit) + i) % 8))
			elif int(byteorder) == 1:
				signalData[int(startByte) + (int(startBit) + i)//8] |= (((int(signalVal, 16) >> i) & 0x01) << ((int(startBit) + i) % 8))
		# retData = [hex(x) for x in signalData]
		retData = signalData

		return retData

	def data_handle(self):
		"""数据处理"""
		pass


	def build_table(self):
		"""创建数组表"""
		pass


# 信号初始值
class PbMsgRevInitVal(PbMsgRevInitDefaultValBase, HexBase):
	"""PB_MsgRevInitVal"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.srcSignalInfoList = []
		self.validDataList = []
		self.PbMsgRevInitValList = []
		self.PB_MsgRevInitVal = []		
		self.structType = ["uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8"] 
		self.structLen = 8
		# self.tableLenAddr = ""
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["PB_MsgRevInitVal"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["PB_MsgRevInitVal"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c0bafa"
		# self.tableLen = 128
		self.tableAddr = self.config.addrInfo["PB_MsgRevInitVal"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["PB_MsgRevInitVal"]["tableLen"])
		self.hexDataList = []

	def data_handle(self):
		"""数据处理"""
		# 创建对应源信号报文数量的初始值列表
		for i in range(len(self.srcSignalInfoList)):
			self.PbMsgRevInitValList.append([0, 0, 0, 0, 0, 0, 0, 0])

		# 轮询全部信号列表数据
		for subList in self.validDataList:
			index = self.srcSignalInfoList.index([subList[self.sigValidHeaderList.index("RxCANID")], subList[self.sigValidHeaderList.index("RxChannel")]])
			# 计算每一行信号数据单元对应的值
			colDataList = self.calc_signa_val(subList[self.sigValidHeaderList.index("inival")],\
												int(subList[self.sigValidHeaderList.index("RxStartBit")])//8,\
												int(subList[self.sigValidHeaderList.index("RxStartBit")])%8,\
												subList[self.sigValidHeaderList.index("RxSigLen")],\
												subList[self.sigValidHeaderList.index("ByteOrder")])
			# 将获得的每一行信号数据单元的值与与初始值列表对应的源信号ID初始值进行位或
			for i in range(8):
				self.PbMsgRevInitValList[index][i] |= colDataList[i]
		
		# 轮询初始值列表，将十进制转换为十六进制；{0:02X}可将十进制数格式化为大写十六进制，并且不足两位前面补0
		for i in range(len(self.PbMsgRevInitValList)):
			self.PbMsgRevInitValList[i] = ["0x{0:02X}u".format(x) for x in self.PbMsgRevInitValList[i]]

			# print(self.PbMsgRevInitValList[i])

	def build_table(self):
		"""创建数组表"""
		self.PB_MsgRevInitVal.append("/*------------------table8: PB_MsgRevInitVal -----------------------------*/\n")
		self.PB_MsgRevInitVal.append("const uint8 PB_MsgRevInitVal[MSG_REV_BUFFER_SIZE] __at(0x00c0bafa)=\n")
		self.PB_MsgRevInitVal.append("{\n")
		for index in range(len(self.srcSignalInfoList)):
			self.PB_MsgRevInitVal.append("/*" + self.srcSignalInfoList[index][0] + "*/")
			self.PB_MsgRevInitVal.append(','.join(self.PbMsgRevInitValList[index]))
			self.PB_MsgRevInitVal.append(",/*NODE_" + get_column_letter(int(self.srcSignalInfoList[index][1])) + " " + str(index) + "*/\n")
		else:
			# 如果列表为空，则以0，填充数组，以防工程编译时不通过.
			if not self.srcSignalInfoList:
				self.PB_MsgRevInitVal.append('\t0,\n')
		self.PB_MsgRevInitVal.append("};")


# 信号失效值
class PbMsgRevDefaultVal(PbMsgRevInitDefaultValBase, HexBase):
	"""PB_MsgRevDefaultVal"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.srcSignalInfoList = []
		self.validDataList = []
		self.PbMsgRevDefaultVal = []
		self.PB_MsgRevDefaultVal = []		
		self.structType = ["uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8", "uint8"] 
		self.structLen = 8
		# self.tableLenAddr = ""
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["PB_MsgRevDefaultVal"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["PB_MsgRevDefaultVal"]["lenType"]
		self.lenHexDataList = []
		# self.tableAddr = "0x00c16c00"
		# self.tableLen = 128
		self.tableAddr = self.config.addrInfo["PB_MsgRevDefaultVal"]["tableAddr"]
		self.tableLen = int(self.config.addrInfo["PB_MsgRevDefaultVal"]["tableLen"])
		self.hexDataList = []

		

		

	def data_handle(self):
		"""数据处理"""
		# 创建对应源信号报文数量的初始值列表
		for i in range(len(self.srcSignalInfoList)):
			self.PbMsgRevDefaultVal.append([0, 0, 0, 0, 0, 0, 0, 0])

		# 轮询全部信号列表数据
		for subList in self.validDataList:
			index = self.srcSignalInfoList.index([subList[self.sigValidHeaderList.index("RxCANID")], subList[self.sigValidHeaderList.index("RxChannel")]])
			# 计算每一行信号数据单元对应的值
			colDataList = self.calc_signa_val(subList[self.sigValidHeaderList.index("dfVal")],\
												int(subList[self.sigValidHeaderList.index("RxStartBit")])//8,\
												int(subList[self.sigValidHeaderList.index("RxStartBit")])%8,\
												subList[self.sigValidHeaderList.index("RxSigLen")],\
												subList[self.sigValidHeaderList.index("ByteOrder")])
			# 将获得的每一行信号数据单元的值与与初始值列表对应的源信号ID初始值进行位或
			for i in range(8):
				self.PbMsgRevDefaultVal[index][i] |= colDataList[i]
		
		# 轮询初始值列表，将十进制转换为十六进制；{0:02X}可将十进制数格式化为大写十六进制，并且不足两位前面补0
		for i in range(len(self.PbMsgRevDefaultVal)):
			self.PbMsgRevDefaultVal[i] = ["0x{0:02X}u".format(x) for x in self.PbMsgRevDefaultVal[i]]

			# print(self.PbMsgRevDefaultVal[i])

	def build_table(self):
		"""创建数组表"""
		self.PB_MsgRevDefaultVal.append("/*------------------table9: PB_MsgRevDefaultVal -----------------------------*/\n")
		self.PB_MsgRevDefaultVal.append("const uint8 SHUGE PB_MsgRevDefaultVal[MSG_REV_BUFFER_SIZE] __at(0x00c16c00)=\n")
		self.PB_MsgRevDefaultVal.append("{\n")
		for index in range(len(self.srcSignalInfoList)):
			self.PB_MsgRevDefaultVal.append("/*" + self.srcSignalInfoList[index][0] + "*/")
			self.PB_MsgRevDefaultVal.append(','.join(self.PbMsgRevDefaultVal[index]))
			self.PB_MsgRevDefaultVal.append(",/*NODE_" + get_column_letter(int(self.srcSignalInfoList[index][1])) + " " + str(index) + "*/\n")
		else:
			# 如果列表为空，则以0，填充数组，以防工程编译时不通过.
			if not self.srcSignalInfoList:
				self.PB_MsgRevDefaultVal.append('\t0,\n')
		self.PB_MsgRevDefaultVal.append("};")


# 报文索引
class Id2IndexTable(RoutingTable, HexBase):
	"""id2index_table"""
	def __init__(self, config):
		"""初始化"""
		self.config = config
		self.msgValidDataList = []
		self.msgValidDataListISR = []
		self.msgDesChList = []
		self.msgDesChListISR = []
		self.signalValidDataList = []
		self.routerTableList = []
		if self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "Qoros_C6M0" or self.config.platInfo == "CHJ":
			self.routerTableListHeader = ["RxCANID", "dest_mo_num", "src_ecu_node", "valid_flg_index", "DLC", "dest_mo_index1", "dest_mo_condition1",\
										  "dest_mo_index2", "dest_mo_condition2", "dest_mo_index3", "dest_mo_condition3", "dest_mo_index4", "dest_mo_condition4",\
										  "dest_mo_index5", "dest_mo_condition5", "cycle", "msg_index", "buf_index", "pre_callback", "post_callback",\
										  "RxInterrupt", "RxDTC", "RxChannel"]
		elif self.config.platInfo == "GAW1.2_OldPlatform":
			self.routerTableListHeader = ["RxCANID", "dest_mo_num", "valid_flg_index", "dest_mo_index1", "dest_mo_condition1",\
									  "dest_mo_index2", "dest_mo_condition2", "dest_mo_index3", "dest_mo_condition3", "dest_mo_index4", "dest_mo_condition4",\
									  "dest_mo_index5", "dest_mo_condition5", "msg_index", "buf_index",\
									  "RxInterrupt", "RxDTC", "RxChannel"]


		self.routerTableISRList = []
		self.routerTableFIFOList = []
		self.PB_MsgRoutingTable = []
		self.idList = []

		self.id2IndexTableA = []
		self.id2IndexTableB = []
		self.id2IndexTableC = []
		self.id2IndexTableD = []
		self.id2IndexTableE = []
		self.id2IndexTableF = []

		self.id2index_table_a = []
		self.id2index_table_b = []
		self.id2index_table_c = []
		self.id2index_table_d = []
		self.id2index_table_e = []
		self.id2index_table_f = []

		self.id2index_table = []
			
		self.structType = ["uint16" for i in range(2048)]
		self.structLen = 4096
		# self.tableLenAddr = ""
		# self.lenType = "uint8"
		self.tableLenAddr = self.config.addrInfo["Id2IndexTableA"]["tableLenAddr"]
		self.lenType = self.config.addrInfo["Id2IndexTableA"]["lenType"]
		self.lenHexDataList = []	
		# self.tableAddrA = "0x00c0a5f6"
		# self.tableAddrB = "0x00c06804"
		# self.tableAddrC = "0x00c0c2ea"
		# self.tableAddrD = "0x00c0d2ea"
		# self.tableAddrE = "0x00c10200"
		# self.tableAddrF = "0x00c11200"
		self.tableAddrA = self.config.addrInfo["Id2IndexTableA"]["tableAddr"]
		self.tableAddrB = self.config.addrInfo["Id2IndexTableB"]["tableAddr"]
		self.tableAddrC = self.config.addrInfo["Id2IndexTableC"]["tableAddr"]
		self.tableAddrD = self.config.addrInfo["Id2IndexTableD"]["tableAddr"]
		self.tableAddrE = self.config.addrInfo["Id2IndexTableE"]["tableAddr"]
		self.tableAddrF = self.config.addrInfo["Id2IndexTableF"]["tableAddr"]
		# self.tableLen = 1
		self.tableLen = int(self.config.addrInfo["Id2IndexTableA"]["tableLen"])
		self.hexDataList = []

	def data_handle(self):
		"""数据处理"""

		super().data_handle()

		for i in range(int("0x7FF", 16) + 1):
			self.id2IndexTableA.append("0xFFFu")
			self.id2IndexTableB.append("0xFFFu")
			self.id2IndexTableC.append("0xFFFu")
			self.id2IndexTableD.append("0xFFFu")
			self.id2IndexTableE.append("0xFFFu")
			self.id2IndexTableF.append("0xFFFu")

		# 将ID单独提取出来，用于后面获取索引
		for subList in self.routerTableFIFOList:
			self.idList.append((subList[self.routerTableListHeader.index("RxCANID")], subList[self.routerTableListHeader.index("RxChannel")]))

		# 轮询所有轮询发送的报文，获得索引值
		for subList in self.routerTableFIFOList:
			if subList[-1] == '1':
				self.id2IndexTableA[int(subList[self.routerTableListHeader.index("RxCANID")], 16)] = \
				"0x{0:04X}u".format(self.idList.index((subList[self.routerTableListHeader.index("RxCANID")], subList[self.routerTableListHeader.index("RxChannel")])))
			elif subList[-1] == '2':
				self.id2IndexTableB[int(subList[self.routerTableListHeader.index("RxCANID")], 16)] = \
				"0x{0:04X}u".format(self.idList.index((subList[self.routerTableListHeader.index("RxCANID")], subList[self.routerTableListHeader.index("RxChannel")])))
			elif subList[-1] == '3':
				self.id2IndexTableC[int(subList[self.routerTableListHeader.index("RxCANID")], 16)] = \
				"0x{0:04X}u".format(self.idList.index((subList[self.routerTableListHeader.index("RxCANID")], subList[self.routerTableListHeader.index("RxChannel")])))
			elif subList[-1] == '4':
				self.id2IndexTableD[int(subList[self.routerTableListHeader.index("RxCANID")], 16)] = \
				"0x{0:04X}u".format(self.idList.index((subList[self.routerTableListHeader.index("RxCANID")], subList[self.routerTableListHeader.index("RxChannel")])))
			elif subList[-1] == '5':
				self.id2IndexTableE[int(subList[self.routerTableListHeader.index("RxCANID")], 16)] = \
				"0x{0:04X}u".format(self.idList.index((subList[self.routerTableListHeader.index("RxCANID")], subList[self.routerTableListHeader.index("RxChannel")])))
			elif subList[-1] == '6':
				self.id2IndexTableF[int(subList[self.routerTableListHeader.index("RxCANID")], 16)] = \
				"0x{0:04X}u".format(self.idList.index((subList[self.routerTableListHeader.index("RxCANID")], subList[self.routerTableListHeader.index("RxChannel")])))

		# print(self.id2IndexTableA)
		# print(self.id2IndexTableB)
		# print(self.id2IndexTableC)
		# print(self.id2IndexTableD)
		# print(self.id2IndexTableE)
		# print(self.id2IndexTableF)

	def build_table(self):
		"""创建数组表"""
		self.id2index_table_a.append("FAR const uint16 id2index_table_a[0x800] __at(0x00c0a5f6) =\n")
		self.id2index_table_b.append("FAR const uint16 id2index_table_b[0x800] __at(0x00c06804) =\n")
		self.id2index_table_c.append("FAR const uint16 id2index_table_c[0x800] __at(0x00c0c2ea) =\n")
		self.id2index_table_d.append("FAR const uint16 id2index_table_d[0x800] __at(0x00c0d2ea) =\n")
		self.id2index_table_e.append("FAR const uint16 id2index_table_e[0x800] __at(0x00c10200) =\n")
		self.id2index_table_f.append("FAR const uint16 id2index_table_f[0x800] __at(0x00c11200) =\n")
		# self.id2index_table_a.append("table8_IdIndex0={\n")
		# self.id2index_table_b.append("table8_IdIndex1={\n")
		# self.id2index_table_c.append("table8_IdIndex2={\n")
		# self.id2index_table_d.append("table8_IdIndex3={\n")
		# self.id2index_table_e.append("table8_IdIndex4={\n")
		# self.id2index_table_f.append("table8_IdIndex5={\n")

		self.id2index_table_a.append("{\n")
		self.id2index_table_b.append("{\n")
		self.id2index_table_c.append("{\n")
		self.id2index_table_d.append("{\n")
		self.id2index_table_e.append("{\n")
		self.id2index_table_f.append("{\n")

		for i in range(int("0x7FF", 16) + 1):
			self.id2index_table_a.append("/*" + "0x{0:03X}".format(i) + "*/" + self.id2IndexTableA[i] + ",\n")
			self.id2index_table_b.append("/*" + "0x{0:03X}".format(i) + "*/" + self.id2IndexTableB[i] + ",\n")
			self.id2index_table_c.append("/*" + "0x{0:03X}".format(i) + "*/" + self.id2IndexTableC[i] + ",\n")
			self.id2index_table_d.append("/*" + "0x{0:03X}".format(i) + "*/" + self.id2IndexTableD[i] + ",\n")
			self.id2index_table_e.append("/*" + "0x{0:03X}".format(i) + "*/" + self.id2IndexTableE[i] + ",\n")
			self.id2index_table_f.append("/*" + "0x{0:03X}".format(i) + "*/" + self.id2IndexTableF[i] + ",\n")

		self.id2index_table_a.append("};\n")
		self.id2index_table_b.append("};\n")
		self.id2index_table_c.append("};\n")
		self.id2index_table_d.append("};\n")
		self.id2index_table_e.append("};\n")
		self.id2index_table_f.append("};\n")

		self.id2index_table.extend(self.id2index_table_a)
		self.id2index_table.extend(self.id2index_table_b)
		self.id2index_table.extend(self.id2index_table_c)
		self.id2index_table.extend(self.id2index_table_d)
		self.id2index_table.extend(self.id2index_table_e)
		self.id2index_table.extend(self.id2index_table_f)


# 诊断请求路由表
class DiagRoutingTable(HexBase):
	"""DiagRoutingTable"""
	pass


if __name__ == '__main__':
	pass
	