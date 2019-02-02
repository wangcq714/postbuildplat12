import re
from copy import deepcopy
from openpyxl.utils import get_column_letter


class CheckError(object):
	'''错误检查'''
	def __init__(self, config):
		self.config = config
		self.id_ch_list = []
		self.first_msg_data_list = []

	def id_is_hex(self, data:str):
		'''判断你是否为16进制'''
		if re.match("^0x(([0-9 A-F]{1,2})|([0-7]{1}[0-9 A-F]{2}))$", data, flags=re.IGNORECASE):
			# print("合法")
			return True
		else:
			# print(data, "不合法")
			return False

	def msg_literal_check(self, msgRoute):
		'''报文数据字面合法性检查'''
		# 深拷贝一份路由数据，不破坏原有数据
		msgDataList = deepcopy(msgRoute.dataList)
		msgHeaderList = deepcopy(msgRoute.headerList)

		tab = "MsgTable"
		row = None # 对应于excel表的行
		col = None # 对应于excel表的列
		hint = "" # ""：没有错误

		ret = False

		# 对报文数据的每一列进行数据校验
		for i in range(len(msgDataList)):
			row = i + 2
			if not self.id_is_hex(msgDataList[i][msgHeaderList.index("TxCANID")]):
				col = get_column_letter(msgHeaderList.index("TxCANID") + 1)
				hint = "发送报文ID格式错误或超出范围，请填写标准的十六进制格式(0x000-0x7FF)"
				break
			if not re.match("^([1-9]{1}[0-9]+)|None$", msgDataList[i][msgHeaderList.index("TxPeriod")]):
				col = get_column_letter(msgHeaderList.index("TxPeriod") + 1)
				hint = "发送报文周期错误（请填写十进制数字且不小于10或NA)"
				break
			if not re.match("^[1-8]{1}$", msgDataList[i][msgHeaderList.index("TxDLC")]):
				col = get_column_letter(msgHeaderList.index("TxDLC") + 1)
				hint = "发送报文长度错误，请填写1-8"
				break
			if not re.match("^[0-6]{1}$", msgDataList[i][msgHeaderList.index("TxChannle")]):
				col = get_column_letter(msgHeaderList.index("TxChannle") + 1)
				hint = "发送报文通道错误，请填写0-6，0代表没有转发关系"
				break

			if not self.id_is_hex(msgDataList[i][msgHeaderList.index("RxCANID")]):
				col = get_column_letter(msgHeaderList.index("RxCANID") + 1)
				hint = "接收报文ID格式错误，请填写标准的十六进制格式(0x000-0x7FF)"
				break
			if not re.match("^([1-9]{1}[0-9]+)|None$", msgDataList[i][msgHeaderList.index("RxPeriod")]):
				col = get_column_letter(msgHeaderList.index("RxPeriod") + 1)
				hint = "接收报文周期错误（填写十进制数字或NA)"
				break
			if not re.match("^[1-8]{1}$", msgDataList[i][msgHeaderList.index("RxDLC")]):
				col = get_column_letter(msgHeaderList.index("RxDLC") + 1)
				hint = "接收报文长度错误，请填写1-8"
				break
			if not re.match("^[1-6]{1}$", msgDataList[i][msgHeaderList.index("RxChannel")]):
				col = get_column_letter(msgHeaderList.index("RxChannel") + 1)
				hint = "接收报文通道错误，请填写1-6"
				break

			if not self.id_is_hex(msgDataList[i][msgHeaderList.index("RxMsk")]):
				col = get_column_letter(msgHeaderList.index("RxMsk") + 1)
				hint = "接收报文掩码错误，请填写标准的十六进制格式(0x000-0xFFF)"
				break
			if not re.match("^[0-1]{1}$", msgDataList[i][msgHeaderList.index("RxInterrupt")]):
				col = get_column_letter(msgHeaderList.index("RxInterrupt") + 1)
				hint = "接收报文是否中断转发填写错误，请填写0或1，0代表不中断接收，1代表中断接收"
				break
			if not re.match("^[YN]{1}$", msgDataList[i][msgHeaderList.index("RxDTC")]):
				col = get_column_letter(msgHeaderList.index("RxDTC") + 1)
				hint = "接收报文是否作为节点丢失DTC判断条件填写错误，请填写Y或N，Y代表是节点丢失报文，N代表不是节点丢失报文"
				break
			if not re.match("^[0-3]{1}$", msgDataList[i][msgHeaderList.index("RouteCondiction")]):
				col = get_column_letter(msgHeaderList.index("RouteCondiction") + 1)
				hint = "接收报文路由条件填写错误，请填写0-3，数字越大限制越小，一般为3"
				break

		
		if hint:
			ret = (tab, row, col, hint)

		return ret

	def signal_literal_check(self, signalRoute):
		'''信号数据字面合法性检查'''
		# 深拷贝一份路由数据，不破坏原有数据
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)

		tab = "SignalTable"
		row = None
		col = None 
		hint = "" # ""：没有错误

		ret = False

		# 对信号数据的每一列进行合法性校验
		for i in range(len(signalDataList)):
			row = i + 2
			if not self.id_is_hex(signalDataList[i][signalHeaderList.index("TxCANID")]):
				col = get_column_letter(signalHeaderList.index("TxCANID") + 1)
				hint = "发送信号ID格式错误，请填写标准的十六进制格式(0x000-0x7FF)"
				break
			if not re.match("^[1-9]{1}[0-9]+$", signalDataList[i][signalHeaderList.index("TxPeriod")]):
				col = get_column_letter(signalHeaderList.index("TxPeriod") + 1)
				hint = "发送信号周期错误（填写十进制数字且大于10)"
				break
			if not re.match("^[1-8]{1}$", signalDataList[i][signalHeaderList.index("TxDLC")]):
				col = get_column_letter(signalHeaderList.index("TxDLC") + 1)
				hint = "发送信号DLC错误，请填写1-8"
				break
			if not re.match("^[1-6]{1}$", signalDataList[i][signalHeaderList.index("TxChannle")]):
				col = get_column_letter(signalHeaderList.index("TxChannle") + 1)
				hint = "发送信号通道错误，请填写1-6"
				break
			if not (re.match("^[0-9]{1,2}$", signalDataList[i][signalHeaderList.index("TxStartBit")]) and (0 <= int(signalDataList[i][signalHeaderList.index("TxStartBit")]) <= 63)):
				col = get_column_letter(signalHeaderList.index("TxStartBit") + 1)
				hint = "发送信号起始位错误，请填写0-63"
				break
			if not (re.match("^[1-9]{1}[0-9]?$", signalDataList[i][signalHeaderList.index("TxSigLen")]) and (1 <= int(signalDataList[i][signalHeaderList.index("TxSigLen")]) <= 64)):
				col = get_column_letter(signalHeaderList.index("TxSigLen") + 1)
				hint = "发送信号长度错误，请填写1-64"
				break
			if self.config.platInfo == "Qoros_C6M0":
				if not re.match("^[0-1]{1}$", signalDataList[i][signalHeaderList.index("TxByteOrder")]):
					col = get_column_letter(signalHeaderList.index("TxByteOrder") + 1)
					hint = "发送字节序错误，请填写0或1，0表示Motorola大端，1表示Intel小端"
					break

			if not self.id_is_hex(signalDataList[i][signalHeaderList.index("RxCANID")]):
				col = get_column_letter(signalHeaderList.index("RxCANID") + 1)
				hint = "接收信号ID格式错误，请填写标准的十六进制格式(0x000-0x7FF)"
				break
			if not re.match("^[1-9]{1}[0-9]+$", signalDataList[i][signalHeaderList.index("RxPeriod")]):
				col = get_column_letter(signalHeaderList.index("RxPeriod") + 1)
				hint = "接收信号周期错误（填写十进制数字且大于10)"
				break
			if not re.match("^[1-8]{1}$", signalDataList[i][signalHeaderList.index("RxDLC")]):
				col = get_column_letter(signalHeaderList.index("RxDLC") + 1)
				hint = "接收信号DLC错误，请填写1-8"
				break
			if not re.match("^[1-6]{1}$", signalDataList[i][signalHeaderList.index("RxChannel")]):
				col = get_column_letter(signalHeaderList.index("RxChannel") + 1)
				hint = "接收信号通道错误，请填写1-6"
				break
			if not (re.match("^[0-9]{1,2}$", signalDataList[i][signalHeaderList.index("RxStartBit")]) and (0 <= int(signalDataList[i][signalHeaderList.index("RxStartBit")]) <= 63)):
				col = get_column_letter(signalHeaderList.index("RxStartBit") + 1)
				hint = "接收信号起始位错误，请填写0-63"
				break
			if not (re.match("^[1-9]{1}[0-9]?$", signalDataList[i][signalHeaderList.index("RxSigLen")]) and (1 <= int(signalDataList[i][signalHeaderList.index("RxSigLen")]) <= 64)):
				col = get_column_letter(signalHeaderList.index("RxSigLen") + 1)
				hint = "接收信号长度错误，请填写1-64"
				break

			if not re.match("^[0-1]{1}$", signalDataList[i][signalHeaderList.index("ByteOrder")]):
				col = get_column_letter(signalHeaderList.index("ByteOrder") + 1)
				hint = "字节序错误，请填写0或1，0表示Motorola大端，1表示Intel小端"
				break
			if not re.match("^[YN]{1}$", signalDataList[i][signalHeaderList.index("RxDTC")]):
				col = get_column_letter(signalHeaderList.index("RxDTC") + 1)
				hint = "接收报文是否作为节点丢失DTC判断条件填写错误，请填写Y或N(大写)，Y代表是节点丢失报文，N代表不是节点丢失报文"
				break
			if not re.match("0x[0-9 A-F]+", signalDataList[i][signalHeaderList.index("inival")], flags=re.IGNORECASE):
				col = get_column_letter(signalHeaderList.index("inival") + 1)
				hint = "接收信号初始值填写错误，请填写标准的十六进制格式"
				break
			if not re.match("0x[0-9 A-F]+", signalDataList[i][signalHeaderList.index("dfVal")], flags=re.IGNORECASE):
				col = get_column_letter(signalHeaderList.index("dfVal") + 1)
				hint = "接收信号失效值填写错误，请填写标准的十六进制格式"
				break



		if hint:
			ret = (tab, row, col, hint)

		return ret

	def msg_logic_check(self, msgRoute):
		'''报文数据逻辑性检查'''
		# 深拷贝一份路由数据，不破坏原有数据
		msgDataList = deepcopy(msgRoute.dataList)
		msgHeaderList = deepcopy(msgRoute.headerList)

		self.id_ch_list = []
		self.first_msg_data_list = []

		tab = "MsgTable"
		row = None # 对应于excel表的行
		col = None # 对应于excel表的列
		hint = "" # ""：没有错误
		ret = False

		# 轮询整个报文数据列表，判断对应发送ID、周期，DLC是否与接收ID、周期、DLC一致，如不一致，则返回报错；
		# 取每一行接收报文ID和对应通道，如果ID和ch第一次出现，则存储在对应列表中；
		# 如不是第一次出现，则判断对应数据是否与第一次出现的数据一致，不一致返回报错。
		for ind, subList in enumerate(msgDataList):
			row = ind +2
			# 进行单行报文数据校验				
			if subList[msgHeaderList.index("TxCANID")] != subList[msgHeaderList.index("RxCANID")]:
				col = get_column_letter(msgHeaderList.index("TxCANID") + 1)
				hint = "发送报文ID与接收报文ID不一致，暂不支持变ID"
				break
			if subList[msgHeaderList.index("TxPeriod")] != subList[msgHeaderList.index("RxPeriod")]:
				col = get_column_letter(msgHeaderList.index("TxPeriod") + 1)
				hint = "发送报文周期与接收报文周期不一致，报文转发暂不支持变周期"
				break
			if subList[msgHeaderList.index("TxDLC")] != subList[msgHeaderList.index("RxDLC")]:
				col = get_column_letter(msgHeaderList.index("TxDLC") + 1)
				hint = "发送报文DLC与接收报文DLC不一致"
				break
			# 前后一致性数据校验
			rx_id_ch = [subList[msgHeaderList.index("RxCANID")], subList[msgHeaderList.index("RxChannel")]]
			if rx_id_ch not in self.id_ch_list:
				self.id_ch_list.append(rx_id_ch)
				self.first_msg_data_list.append(subList)			
			else:
				first_msg_data = self.first_msg_data_list[self.id_ch_list.index(rx_id_ch)]
				# 逐列进行校验
				if subList[msgHeaderList.index("TxCANID")] != first_msg_data[msgHeaderList.index("TxCANID")]:
					col = get_column_letter(msgHeaderList.index("TxCANID") + 1)
					hint = "发送报文ID与与该行之上同网段同ID接收报文对应发送报文ID不一致，暂不支持变ID"
					break
				if subList[msgHeaderList.index("TxPeriod")] != first_msg_data[msgHeaderList.index("TxPeriod")]:
					col = get_column_letter(msgHeaderList.index("TxPeriod") + 1)
					hint = "发送报文周期与该行之上同网段同ID接收报文对应发送报文周期不一致，报文转发暂不支持变周期"
					break
				if subList[msgHeaderList.index("TxDLC")] != first_msg_data[msgHeaderList.index("TxDLC")]:
					col = get_column_letter(msgHeaderList.index("TxDLC") + 1)
					hint = "发送报文DLC与该行之上同网段同ID接收报文对应发送报文DLC不一致"
					break
				if subList[msgHeaderList.index("RxPeriod")] != first_msg_data[msgHeaderList.index("RxPeriod")]:
					col = get_column_letter(msgHeaderList.index("RxPeriod") + 1)
					hint = "接收报文周期与该行之上同网段同ID报文周期不一致"
					break
				if subList[msgHeaderList.index("RxDLC")] != first_msg_data[msgHeaderList.index("RxDLC")]:
					col = get_column_letter(msgHeaderList.index("RxDLC") + 1)
					hint = "接收报文DLC与该行之上同网段同ID报文DLC不一致"
					break
				if subList[msgHeaderList.index("RxMsk")] != first_msg_data[msgHeaderList.index("RxMsk")]:
					col = get_column_letter(msgHeaderList.index("RxMsk") + 1)
					hint = "接收报文掩码与该行之上同网段同ID报文掩码不一致"
					break
				if subList[msgHeaderList.index("RxInterrupt")] != first_msg_data[msgHeaderList.index("RxInterrupt")]:
					col = get_column_letter(msgHeaderList.index("RxInterrupt") + 1)
					hint = "接收报文是否中断配置与该行之上同网段同ID报文是否中断配置不一致"
					break
				if subList[msgHeaderList.index("RxDTC")] != first_msg_data[msgHeaderList.index("RxDTC")]:
					col = get_column_letter(msgHeaderList.index("RxDTC") + 1)
					hint = "接收报文节点丢失DTC配置与该行之上同网段同ID报文节点丢失DTC配置不一致"
					break
				if subList[msgHeaderList.index("RouteCondiction")] != first_msg_data[msgHeaderList.index("RouteCondiction")]:
					col = get_column_letter(msgHeaderList.index("RouteCondiction") + 1)
					hint = "接收报文路由条件配置与该行之上同网段同ID报文路由条件配置不一致"
					break

		if hint:
			ret = (tab, row, col, hint)

		return ret

	def signal_logic_check(self, signalRoute):
		'''信号数据逻辑性检查'''
		# 深拷贝一份路由数据，不破坏原有数据
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)

		rx_id_ch_list = []
		first_rxsignal_data_list = []
		tx_id_ch_list = []
		first_txsignal_data_list = []

		tab = "SignalTable"
		row = None # 对应于excel表的行
		col = None # 对应于excel表的列
		hint = "" # ""：没有错误
		ret = False

		# 将信号分为接收信号和发送信号分别处理；
		# 轮询每一个信号，对信号数据进行校验；
		# 提取ID和CH，如果ID和CH是第一次出现，将其放入列表中；
		# 如果ID和CH已经出现过，则将当前信号与已出现的信号数据进行校验；
		for ind, subList in enumerate(signalDataList):
			row = ind +2
			# 接收信号数据校验			
			# 单行数据校验
			if subList[signalHeaderList.index("ByteOrder")] == '0':
				if not (int(subList[signalHeaderList.index("RxSigLen")]) <= \
					(int(subList[signalHeaderList.index("RxStartBit")]) // 8) * 8 + (8 - (int(subList[signalHeaderList.index("RxStartBit")]) % 8))):
					col = get_column_letter(signalHeaderList.index("RxSigLen") + 1)
					hint = "接收信号长度超限"
					break
			elif subList[signalHeaderList.index("ByteOrder")] == '1':
				if not ((int(subList[signalHeaderList.index("RxStartBit")]) + int(subList[signalHeaderList.index("RxSigLen")]) - 1) <= 63):
					col = get_column_letter(signalHeaderList.index("RxSigLen") + 1)
					hint = "接收信号长度超限"
					break
			if not (int(subList[signalHeaderList.index("inival")], 16) < 2**int(subList[signalHeaderList.index("RxSigLen")])):
				col = get_column_letter(signalHeaderList.index("inival") + 1)
				hint = "接收信号初始值超出合理值范围"
				break
			if not (int(subList[signalHeaderList.index("dfVal")], 16) < 2**int(subList[signalHeaderList.index("RxSigLen")])):
				col = get_column_letter(signalHeaderList.index("dfVal") + 1)
				hint = "接收信号失效值超出合理值范围"
				break
			if not (int(subList[signalHeaderList.index("RxSigLen")]) == int(subList[signalHeaderList.index("TxSigLen")])):
				col = get_column_letter(signalHeaderList.index("RxSigLen") + 1)
				hint = "接收信号长度与发送信号长度不一致"
				break

			# 数据前后一致性校验
			rx_id_ch = [subList[signalHeaderList.index("RxCANID")], subList[signalHeaderList.index("RxChannel")]]
			if rx_id_ch not in rx_id_ch_list:
				rx_id_ch_list.append(rx_id_ch)
				first_rxsignal_data_list.append(subList)
			else:
				# 逐列进行校验
				first_rxsignal_data = first_rxsignal_data_list[rx_id_ch_list.index(rx_id_ch)]
				if subList[signalHeaderList.index("RxPeriod")] != first_rxsignal_data[signalHeaderList.index("RxPeriod")]:
					col = get_column_letter(signalHeaderList.index("RxPeriod") + 1)
					hint = "接收信号周期与该行之上同网段同ID信号周期不一致"
					break
				if subList[signalHeaderList.index("RxDLC")] != first_rxsignal_data[signalHeaderList.index("RxDLC")]:
					col = get_column_letter(signalHeaderList.index("RxDLC") + 1)
					hint = "接收信号DLC与该行之上同网段同ID信号DLC不一致"
					break
				if subList[signalHeaderList.index("ByteOrder")] != first_rxsignal_data[signalHeaderList.index("ByteOrder")]:
					col = get_column_letter(signalHeaderList.index("ByteOrder") + 1)
					hint = "接收信号字节序与该行之上同网段同ID信号字节序不一致"
					break
				if subList[signalHeaderList.index("RxDTC")] != first_rxsignal_data[signalHeaderList.index("RxDTC")]:
					col = get_column_letter(signalHeaderList.index("RxDTC") + 1)
					hint = "接收信号节点丢失DTC配置与该行之上同网段同ID信号节点丢失DTC配置不一致"
					break
			# 发送信号数据校验
			# 单行数据校验,新平台、老平台表一致；观致增加了发送字节序特殊另行elif校验
			if self.config.platInfo == "GAW1.2_OldPlatform" or self.config.platInfo == "GAW1.2_NewPlatform":
				if subList[signalHeaderList.index("ByteOrder")] == '0':
					if not (int(subList[signalHeaderList.index("TxSigLen")]) <= \
						(int(subList[signalHeaderList.index("TxStartBit")]) // 8) * 8 + (8 - (int(subList[signalHeaderList.index("TxStartBit")]) % 8))):
						col = get_column_letter(signalHeaderList.index("TxSigLen") + 1)
						hint = "发送信号长度超限"
						break
				elif subList[signalHeaderList.index("ByteOrder")] == '1':
					if not ((int(subList[signalHeaderList.index("TxStartBit")]) + int(subList[signalHeaderList.index("TxSigLen")]) - 1) <= 63):
						col = get_column_letter(signalHeaderList.index("TxSigLen") + 1)
						hint = "发送信号长度超限"
						break
				if not (int(subList[signalHeaderList.index("TxSigLen")]) == int(subList[signalHeaderList.index("RxSigLen")])):
					col = get_column_letter(signalHeaderList.index("TxSigLen") + 1)
					hint = "发送信号长度与接收信号长度不一致"
					break

				tx_id_ch = [subList[signalHeaderList.index("TxCANID")], subList[signalHeaderList.index("TxChannle")]]	
				if tx_id_ch not in tx_id_ch_list:
					tx_id_ch_list.append(tx_id_ch)
					first_txsignal_data_list.append(subList)

				else:
					# 逐列进行数据校验
					first_txsignal_data = first_txsignal_data_list[tx_id_ch_list.index(tx_id_ch)]
					if subList[signalHeaderList.index("TxPeriod")] != first_txsignal_data[signalHeaderList.index("TxPeriod")]:
						col = get_column_letter(signalHeaderList.index("TxPeriod") + 1)
						hint = "发送信号周期与该行之上同网段同ID信号周期不一致"
						break
					if subList[signalHeaderList.index("TxDLC")] != first_txsignal_data[signalHeaderList.index("TxDLC")]:
						col = get_column_letter(signalHeaderList.index("TxDLC") + 1)
						hint = "发送信号DLC与该行之上同网段同ID信号DLC不一致"
						break
					if subList[signalHeaderList.index("ByteOrder")] != first_txsignal_data[signalHeaderList.index("ByteOrder")]:
						col = get_column_letter(signalHeaderList.index("ByteOrder") + 1)
						hint = "发送信号字节序与该行之上同网段同ID信号字节序不一致"
						break
			elif self.config.platInfo == "Qoros_C6M0":
				if subList[signalHeaderList.index("TxByteOrder")] == '0':
					if not (int(subList[signalHeaderList.index("TxSigLen")]) <= \
						(int(subList[signalHeaderList.index("TxStartBit")]) // 8) * 8 + (8 - (int(subList[signalHeaderList.index("TxStartBit")]) % 8))):
						col = get_column_letter(signalHeaderList.index("TxSigLen") + 1)
						hint = "发送信号长度超限"
						break
				elif subList[signalHeaderList.index("TxByteOrder")] == '1':
					if not ((int(subList[signalHeaderList.index("TxStartBit")]) + int(subList[signalHeaderList.index("TxSigLen")]) - 1) <= 63):
						col = get_column_letter(signalHeaderList.index("TxSigLen") + 1)
						hint = "发送信号长度超限"
						break
				if not (int(subList[signalHeaderList.index("TxSigLen")]) == int(subList[signalHeaderList.index("RxSigLen")])):
					col = get_column_letter(signalHeaderList.index("TxSigLen") + 1)
					hint = "发送信号长度与接收信号长度不一致"
					break

				tx_id_ch = [subList[signalHeaderList.index("TxCANID")], subList[signalHeaderList.index("TxChannle")]]	
				if tx_id_ch not in tx_id_ch_list:
					tx_id_ch_list.append(tx_id_ch)
					first_txsignal_data_list.append(subList)

				else:
					# 逐列进行数据校验
					first_txsignal_data = first_txsignal_data_list[tx_id_ch_list.index(tx_id_ch)]
					if subList[signalHeaderList.index("TxPeriod")] != first_txsignal_data[signalHeaderList.index("TxPeriod")]:
						col = get_column_letter(signalHeaderList.index("TxPeriod") + 1)
						hint = "发送信号周期与该行之上同网段同ID信号周期不一致"
						break
					if subList[signalHeaderList.index("TxDLC")] != first_txsignal_data[signalHeaderList.index("TxDLC")]:
						col = get_column_letter(signalHeaderList.index("TxDLC") + 1)
						hint = "发送信号DLC与该行之上同网段同ID信号DLC不一致"
						break
					if subList[signalHeaderList.index("TxByteOrder")] != first_txsignal_data[signalHeaderList.index("TxByteOrder")]:
						col = get_column_letter(signalHeaderList.index("TxByteOrder") + 1)
						hint = "发送信号字节序与该行之上同网段同ID信号字节序不一致"
						break

		if hint:
			ret = (tab, row, col, hint)

		return ret

	def msgsignal_logic_check(self, msgRoute, signalRoute):
		'''报文和信号混合逻辑检查'''
		# 深拷贝一份路由数据，不破坏原有数据
		msgDataList = deepcopy(msgRoute.dataList)
		msgHeaderList = deepcopy(msgRoute.headerList)
		signalDataList = deepcopy(signalRoute.dataList)
		signalHeaderList = deepcopy(signalRoute.headerList)

		tab = "SignalTable"
		row = None # 对应于excel表的行
		col = None # 对应于excel表的列
		hint = "" # ""：没有错误
		ret = False

		# 在报文数据校验通过的前提下，轮询信号数据，校验报文和信号数据的一致性
		for ind, subList in enumerate(signalDataList):
			row = ind +2
			rx_id_ch = [subList[signalHeaderList.index("RxCANID")], subList[signalHeaderList.index("RxChannel")]]
			# 如果同网段同ID既做报文又做源信号，则进行报文和源信号的周期、DLC、和DTC配置的一致性校验
			if rx_id_ch in self.id_ch_list:
				if not (subList[signalHeaderList.index("RxPeriod")] == self.first_msg_data_list[self.id_ch_list.index(rx_id_ch)][msgHeaderList.index("RxPeriod")]):
					col = get_column_letter(signalHeaderList.index("RxPeriod") + 1)
					hint = "信号表中的接收信号周期与报文表中(第" + str(msgDataList.index(self.first_msg_data_list[self.id_ch_list.index(rx_id_ch)]) + 2) + \
							"行第"+ get_column_letter(msgHeaderList.index("RxPeriod") + 1) + "列)" + "同网段同ID报文周期不一致"
					break
				if not (subList[signalHeaderList.index("RxDLC")] == self.first_msg_data_list[self.id_ch_list.index(rx_id_ch)][msgHeaderList.index("RxDLC")]):
					col = get_column_letter(signalHeaderList.index("RxDLC") + 1)
					hint = "信号表中的接收信号DLC与报文表中(第" + str(msgDataList.index(self.first_msg_data_list[self.id_ch_list.index(rx_id_ch)]) + 2) + \
							"行第"+ get_column_letter(msgHeaderList.index("RxDLC") + 1) + "列)" + "同网段同ID报文DLC不一致"
					break
				if not (subList[signalHeaderList.index("RxDTC")] == self.first_msg_data_list[self.id_ch_list.index(rx_id_ch)][msgHeaderList.index("RxDTC")]):
					col = get_column_letter(signalHeaderList.index("RxDTC") + 1)
					hint = "信号表中的接收信号节点丢失DTC配置与报文表中(第" + str(msgDataList.index(self.first_msg_data_list[self.id_ch_list.index(rx_id_ch)]) + 2) + \
							"行第"+ get_column_letter(msgHeaderList.index("RxDTC") + 1) + "列)" + "同网段同ID报文节点丢失DTC配置不一致"
					break

		if hint:
			ret = (tab, row, col, hint)

		return ret


if __name__ == '__main__':
	checkError = CheckError()
	for i in range(int("0x800", 16)):
		checkError.id_is_hex(hex(i))
	# checkError.is_dec("None")
