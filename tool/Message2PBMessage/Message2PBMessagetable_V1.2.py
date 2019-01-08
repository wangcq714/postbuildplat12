from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
import csv
from tkinter import *
from tkinter.filedialog import askdirectory,askopenfilename


#获取通道映射
def get_channalmapping(sheet):
	for i in range(6):
		ChannalMapping[sheet['N' + str(i + 3)].value] = sheet['M' + str(i + 3)].value

#读取指定行数据
def read_line(sheet, column) -> list:
	src_column_data = []
	for cell in list(sheet.rows)[column][:column_index_from_string('O'):]:
		src_column_data.append(str(cell.value).rstrip('\n'))
	#print(src_row_data)
	return src_column_data

#读取表中所有有效行数据
def read_lines_all(sheet):
	for i in range(2, sheet.max_row):
		tmpList = read_line(sheet, i)
		if tmpList[0] != "None" and tmpList[column_index_from_string('F') - 1] != "DCAN":
			src_row_data_all.append(tmpList)
		else:
			break

#创建目标单行数据
def build_des_column_data(LineNumber, Txchannal, src_num) -> list:
	global src_row_data_all
	global ChannalMapping
	des_column_data_list = []

	des_column_data_list.append(str(LineNumber))
	des_column_data_list.append(src_row_data_all[src_num][column_index_from_string('B') - 1])
	des_column_data_list.append(src_row_data_all[src_num][column_index_from_string('A') - 1])
	des_column_data_list.append(str(src_row_data_all[src_num][column_index_from_string('C') - 1]))
	des_column_data_list.append(str(src_row_data_all[src_num][column_index_from_string('D') - 1]))
	des_column_data_list.append(Txchannal)
	des_column_data_list.append(src_row_data_all[src_num][column_index_from_string('B') - 1])
	des_column_data_list.append(src_row_data_all[src_num][column_index_from_string('A') - 1])
	des_column_data_list.append(str(src_row_data_all[src_num][column_index_from_string('C') - 1]))
	des_column_data_list.append(str(src_row_data_all[src_num][column_index_from_string('D') - 1]))
	des_column_data_list.append(Can2num[ChannalMapping[src_row_data_all[src_num][column_index_from_string('F') - 1]]])
	des_column_data_list.append("0x7FF")
	if src_row_data_all[src_num][column_index_from_string('O') - 1] == "Y":
		des_column_data_list.append('1')
	else:
		des_column_data_list.append('0')
	if src_row_data_all[src_num][column_index_from_string('E') - 1] == "None" or src_row_data_all[src_num][column_index_from_string('E') - 1] == "NA":
		des_column_data_list.append('N')
	else:
		des_column_data_list.append('Y')
	des_column_data_list.append('3')

	return des_column_data_list

#获取要发送的通道
def get_TxChannal(src_num) -> list:
	global src_row_data_all
	TxChannalList = []
	i = 1
	for tick in src_row_data_all[src_num][column_index_from_string('L') - 1 : column_index_from_string('G') -2 : -1]:
	#print(src_row_data_all[src_num][column_index_from_string('F')])
		if tick == '√':
			TxChannalList.append(i)
		i += 1

	return TxChannalList

#创建所有目标数据
def build_des_column_data_all():
	LineNumber = 1
	for num in range(len(src_row_data_all)):
		TxChList = get_TxChannal(num)
		#print(TxChList)
		if len(TxChList) > 0:
			for TxCh in TxChList:
				des_column_data_all.append(build_des_column_data(LineNumber, TxCh, num))
				LineNumber += 1
		else:
			des_column_data_all.append(build_des_column_data(LineNumber, 0, num))
			LineNumber += 1

#主函数
def main():

	# 默认可读写，若有需要可以指定write_only和read_only为True
	wb = load_workbook(pathname)

	# 根据sheet名字获得sheet
	sheet = wb['Sheet1']

	print(sheet.max_row)

	#通道映射
	get_channalmapping(sheet)
	print(ChannalMapping)

	#读取原表中的所有数据
	read_lines_all(sheet)

	#创建目标表列表
	build_des_column_data_all()

	#将目标列表写入目标文件中
	with open(pathname[:pathname.rfind('/')+1] + "MessageRoute.csv", 'w', newline='') as csvfile:	
		writer = csv.writer(csvfile)
		#写入表头
		writer.writerow(TableHeader)
		#写入目标数据
		writer.writerows(des_column_data_all)

	#print(src_row_data_all)
	#print(des_column_data_all)

	print("------------------Finished--------------------------")

src_row_data_all = []   #存取源表的所有行数据列表
des_column_data_all = []   #存取目标表的所有行数据列表
ChannalMapping = {}        #CAN通道与系统框图通道映射
Can2num = {"CAN1":1, "CAN2":2, "CAN3":3, "CAN4":4, "CAN5":5, "CAN6":6}  #CAN通道与数字映射
TableHeader = ["LineNumber", "TxMessageName", "TxCANID", "TxPeriod", "TxDLC", "TxChannle", "RxMessageName", \
				"RxCANID", "RxPeriod", "RxDLC", "RxChannel", "RxMsk", "RxInterrupt", "RxDTC", "RouteCondiction"]
#读取文件路径名
pathname = askopenfilename(filetypes = [("Excel",".xlsx")])

if pathname != '':
	main()

