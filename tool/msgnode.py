import os
import time
# from tkinter import *
from tkinter.filedialog import askdirectory,askopenfilename
import functools
from multiprocessing import Pool, Process
from threading import Thread, Lock

# from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
# from openpyxl import Workbook
# import csv

from pandas import read_excel


#计算函数运行时间装饰器
def calc_runtime(func):
	@functools.wraps(func)
	def wrapper(*args, **kwargs):
		start_time = time.time()
		ret = func(*args, **kwargs)
		stop_time = time.time()
		print(func.__name__ + "运行时间： %0.3fs"%(stop_time - start_time))
		return ret
	return wrapper


class MsgNodeEnable(object):
	'''报文节点使能配置'''
	def __init__(self):
		self.pathname = ""
		self.CanNode_data_all = []   #存取源表的所有行数据列表
		self.src_row_data_all = []   #存取源表的所有行数据列表
		self.src_row_data_polling = [] #存放轮询ID
		self.src_row_data_polling_sorted = [] #存放排序过轮询ID
		self.src_row_data_ISR = []#存放中断ID
		self.src_row_data_ISR_sorted = []#存放排序过中断ID

	def get_file_pathname(self):
		"""获取路由表路径"""
		self.pathname = askopenfilename(filetypes = [("Excel",".xlsx")])
		if self.pathname == ():
			self.pathname = ""
		# print(self.self.self.pathname)
		# print(type(self.self.self.pathname))

	#将读取的源数据按轮询和中断转发分别存储
	#@calc_runtime
	def split_src_row_data_all(self, SourceDataList):
		PB_ISR_ID_Flag_list = []
		PB_Polling_ID_Flag_list = []
		for ListValue in SourceDataList:
			if ListValue[column_index_from_string('O') - 1] == 'Y':
				if ListValue[column_index_from_string('A') - 1].strip(' ') not in PB_ISR_ID_Flag_list:
					PB_ISR_ID_Flag_list.append(ListValue[column_index_from_string('A') - 1].strip(' '))
					self.src_row_data_ISR.append(ListValue)
			else:
				if ListValue[column_index_from_string('A') - 1].strip(' ') not in PB_Polling_ID_Flag_list:
					PB_Polling_ID_Flag_list.append(ListValue[column_index_from_string('A') - 1].strip(' '))
				self.src_row_data_polling.append(ListValue)

	#依照ID大小对读出的节点丢失行进行排序
	#@calc_runtime
	def sort_src_row_data_all_Missing(self, sourcelist) -> list:
		sort_src_row_data_all_list = []
		i = 0
		for row in sourcelist:
			if row[0] != "NA" and row[0] != "None":
				if i == 0:
					sort_src_row_data_all_list.append(row)
				else:
					for j in range(len(sort_src_row_data_all_list)):
						if int(row[0].strip(' '), 16) < int(sort_src_row_data_all_list[j][0].strip(' '), 16):
							sort_src_row_data_all_list.insert(j, row)
							break
						else:
							if j == len(sort_src_row_data_all_list) - 1:
								sort_src_row_data_all_list.append(row)
				i += 1
		return sort_src_row_data_all_list

	#创建中断报文节点表
	#@calc_runtime
	def build_PB_MsgNodeTable_ISR(self, sourcelist) -> list:
		PB_MsgNodeTable_list = []
		PB_MsgNodeTable_str = "const PB_TABLE_1_Auxiliary PB_MsgNodeTable_ISR[TABLE_1_SIZE] =" + '\n'
		PB_MsgNodeTable_str += '{' + '\n'
		for ListValue in sourcelist:
			if ListValue[column_index_from_string('P') - 1] != "NA" and ListValue[column_index_from_string('P') - 1] != "None":
				PB_MsgNodeTable_str += '\t' + "NODE_" + ListValue[column_index_from_string('P') - 1] + ',' + '\n'
			else:
				PB_MsgNodeTable_str += '\t' + str(255) + ',' + '\n'
		PB_MsgNodeTable_str += "};"
		PB_MsgNodeTable_list.append(PB_MsgNodeTable_str)

		return PB_MsgNodeTable_list

	#创建轮询报文节点表
	#@calc_runtime
	def build_PB_MsgNodeTable_Polling(self, sourcelist) -> list:
		PB_MsgNodeTable_list = []
		PB_MsgNodeTable_str = "const PB_TABLE_1_Auxiliary SHUGE PB_MsgNodeTable_Polling[TABLE_1_SIZE] =" + '\n'
		PB_MsgNodeTable_str += '{' + '\n'
		for ListValue in sourcelist:
			if ListValue[column_index_from_string('P') - 1] != "NA" and ListValue[column_index_from_string('P') - 1] != "None":
				PB_MsgNodeTable_str += '\t' + "NODE_" + ListValue[column_index_from_string('P') - 1] + ',' + '\n'
			else:
				PB_MsgNodeTable_str += '\t' + str(255) + ',' + '\n'
		PB_MsgNodeTable_str += "};"
		PB_MsgNodeTable_list.append(PB_MsgNodeTable_str)

		return PB_MsgNodeTable_list

	#计算EEP_ConfigVal
	def cal_EEP_ConfigVal(self, ListValue) -> str:
		EEP_ConfigVal = int(ListValue[column_index_from_string('B') - 1]) * 8 + int(ListValue[column_index_from_string('C') - 1])
		if EEP_ConfigVal < 10:
			return ' ' + str(EEP_ConfigVal)
		else:
			return str(EEP_ConfigVal)

	#创建gDiag_ECUNode_Cfg_array元素
	def build_gDiag_ECUNode_Cfg_array_element(self, ListValue) -> str:
		gDiag_ECUNode_Cfg_array_element_str = '\t' + '{'
		gDiag_ECUNode_Cfg_array_element_str += "NODE_" + ListValue[0] + ',' + '\t'*2
		gDiag_ECUNode_Cfg_array_element_str += self.cal_EEP_ConfigVal(ListValue) + "},"
		gDiag_ECUNode_Cfg_array_element_str += '\n'

		return gDiag_ECUNode_Cfg_array_element_str

	#创建gDiag_ECUNode_Cfg_array

	def build_gDiag_ECUNode_Cfg_array(self, SourceDataList) -> list:
		gDiag_ECUNode_Cfg_list = []
		gDiag_ECUNode_Cfg_str = "const TYPE_DIAG_NODE_CFG_STRUCT gDiag_ECUNode_Cfg[NODE_CFG_MAX] =" + '\n'
		gDiag_ECUNode_Cfg_str += '{' + '\n'
		for ListValue in SourceDataList:
			gDiag_ECUNode_Cfg_str += self.build_gDiag_ECUNode_Cfg_array_element(ListValue)
		gDiag_ECUNode_Cfg_str += "};" + '\n'
		gDiag_ECUNode_Cfg_list.append(gDiag_ECUNode_Cfg_str)
		return gDiag_ECUNode_Cfg_list

	def mkdir(self, path:str): 
	    # 去除首位空格、尾部\
	    path=path.strip()
	    # 判断结果
	    if not os.path.exists(path):
	        os.makedirs(path) 

	#写文件
	@calc_runtime
	def write_file(self):
		self.mutex.acquire()
		
		self.mkdir("output")
		# with open(self.pathname[:self.pathname.rfind('/')+1] + "MsgNodeTable.c",'w') as Cfile:
		with open("output/MsgNodeTable.c",'w') as Cfile:
			for tmp in self.build_gDiag_ECUNode_Cfg_array(self.CanNode_data_all):
				Cfile.write(tmp)

			Cfile.write('\n'*2)
			#将gDEM_taDTCCfgPattern数组写入
			#print(self.src_row_data_ISR_sorted)
			for tmp in self.build_PB_MsgNodeTable_ISR(self.src_row_data_ISR_sorted):
				Cfile.write(tmp)

			Cfile.write('\n'*2)
			#print(self.src_row_data_polling_sorted)
			for tmp in self.build_PB_MsgNodeTable_Polling(self.src_row_data_polling_sorted):
				Cfile.write(tmp)
		self.mutex.release()

	#主函数
	@calc_runtime
	def nodeconfig_main(self):

		# # 根据sheet名字获得sheet
		# sheet = wb['CANNode']
		# print(sheet.max_row)

		# #读取原表中的所有数据
		# self.CanNode_data_all = read_lines_all(sheet, 'C')
		# #print(self.src_row_data_all)

		dataFrame = read_excel(self.pathname, sheet_name="CANNode", header=None, na_values="", usecols="A:C")
		# print(dataFrame)
		# 将Frame格式数据转换成列表
		self.CanNode_data_all = dataFrame.fillna("None").values[2:].tolist()	
		# print(self.self.src_row_data_all)

		#print(os.getpid())
		print("------------------nodeconfig_main  Finished--------------------------")

	#主函数
	@calc_runtime
	def id2nodeindex_main(self):
		# # 根据sheet名字获得sheet
		# sheet = wb['Sheet1']
		# print(sheet.max_row)

		# #读取原表中的所有数据
		# self.src_row_data_all = read_lines_all(sheet, 'F')
		# #print(self.src_row_data_all)

		dataFrame = read_excel(self.pathname, sheet_name="Sheet1", header=None, na_values="", usecols="A:Q")
		# print(dataFrame)
		# 将Frame格式数据转换成列表
		self.src_row_data_all = dataFrame.fillna("None").values[2:].tolist()	
		# print(self.src_row_data_all)

		self.src_row_data_all = [subList for subList in self.src_row_data_all if subList[column_index_from_string('Q') - 1] != 'Y']
		# print(self.src_row_data_all)

		#以轮询和中断拆分数据
		self.split_src_row_data_all(self.src_row_data_all)
		#print(self.src_row_data_ISR)
		#print("----------------------华丽的分割线---------------------")
		#print(self.src_row_data_polling)

		#以ID大小排序
		self.src_row_data_ISR_sorted = self.sort_src_row_data_all_Missing(self.src_row_data_ISR)
		#print(self.src_row_data_ISR_sorted)
		self.src_row_data_polling_sorted = self.sort_src_row_data_all_Missing(self.src_row_data_polling)
		#print(self.src_row_data_polling_sorted)
		# print(self.src_row_data_ISR_sorted)
		# print("----------------------华丽的分割线---------------------")
		# print(self.src_row_data_polling_sorted)
		#print(len(self.src_row_data_polling_sorted))
		
		#print(os.getpid())
		print("------------------id2nodeindex_main Finished--------------------------")

	def main(self):
		self.mutex = Lock()

		Th_nodeconfig_main =Thread(target=self.nodeconfig_main, args=())
		Th_nodeconfig_main.start()
		#print(Th_nodeconfig_main.name)

		Th_id2nodeindex_main =Thread(target=self.id2nodeindex_main, args=())
		Th_id2nodeindex_main.start()
		#print(Th_id2nodeindex_main.name)
		Th_nodeconfig_main.join()
		Th_id2nodeindex_main.join()


		#nodeconfig_main()
		#id2nodeindex_main()

		self.write_file()

		print("------------------END------------------------------")


if __name__ == "__main__":
	msgNodeEnable = MsgNodeEnable()
	msgNodeEnable.get_file_pathname()
	msgNodeEnable.main()

		
