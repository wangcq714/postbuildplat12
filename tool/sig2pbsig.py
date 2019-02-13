# V1.1

import os
# import pandas
from pandas import read_excel
# from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
# from openpyxl import Workbook
import csv
from tkinter.filedialog import askdirectory, askopenfilename



class SignalTableConvert(object):
	'''信号表转换'''
	def __init__(self, config):
		#通道映射
		self.config = config
		self.Can2num = {"CAN1":1, "CAN2":2, "CAN3":3, "CAN4":4, "CAN5":5, "CAN6":6}
		#目标Excel表头
		if self.config.platInfo == "GAW1.2_OldPlatform" or self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "CHJ":
			self.TableHeader = ["SignalName", "TxMeesageName", "TxCANID", "TxPeriod", "TxDLC", "TxChannle", \
							"TxStartBit", "TxSigLen", "RxMeesageName", "RxCANID", "RxPeriod", "RxDLC", \
							"RxChannel", "RxStartBit", "RxSigLen", "ByteOrder", "RxDTC", "inival", "dfVal", "desName"]
		elif self.config.platInfo == "Qoros_C6M0":
			self.TableHeader = ["SignalName", "TxMeesageName", "TxCANID", "TxPeriod", "TxDLC", "TxChannle", \
							"TxStartBit", "TxSigLen", "TxByteOrder", "RxMeesageName", "RxCANID", "RxPeriod", "RxDLC", \
							"RxChannel", "RxStartBit", "RxSigLen", "ByteOrder", "RxDTC", "inival", "dfVal", "desName"]
		self.pathname = ""
		self.ChannalMapping = {}
		self.src_table = {}
		self.des_table = []

	def get_file_pathname(self):
		"""获取路由表路径"""
		self.pathname = askopenfilename(filetypes = [("Excel",".xlsx")])
		if self.pathname == ():
			self.pathname = ""
		# print(self.self.pathname)
		# print(type(self.self.pathname))

	# #获取通道映射
	# def get_channalmapping(self, sheet):
	# 	for i in range(6):
	# 		self.ChannalMapping[sheet['W' + str(i + 3)].value] = sheet['V' + str(i + 3)].value

	#通道映射,如果dir==0,则为CAN1：netname,dir == 1, netname:dir
	def getChannalMapping_pandas(self, dataFrame, dir) -> dict:
		self.ChannalMapping = {}
		if dir == 0:
			for i in range(6):
				# self.ChannalMapping[dataFrame.values[i + 2][column_index_from_string('N') - 1]] = dataFrame.values[i +2][column_index_from_string('M') - 1].lower()
				self.ChannalMapping[dataFrame.values[i + 2][1]] = dataFrame.values[i +2][0]
		elif dir == 1:
			for i in range(6):
				# self.ChannalMapping[dataFrame.values[i + 2][column_index_from_string('M') - 1]] = dataFrame.values[i +2][column_index_from_string('N') - 1]
				self.ChannalMapping[dataFrame.values[i + 2][0]] = dataFrame.values[i +2][1]
		return self.ChannalMapping

	#获取发送报文的名字
	def get_TxMeesageName(self) -> list:
		TxMeesageName = []
		for i in range(len(self.src_table["des_目标网段ID"])):
			TxMeesageName.append(self.src_table["des_目标网段"][i] + '_' + self.src_table["des_目标网段ID"][i][2:]) 
		return TxMeesageName

	#获取接收报文的名字
	def get_RxMeesageName(self) -> list:
		RxMeesageName = []
		for i in range(len(self.src_table["src_源报文ID"])):
			RxMeesageName.append("GW_" + self.src_table["src_源报文ID"][i][2:]) 
		return RxMeesageName

	#获取发送至目标网段的数字代号
	def ToNetwork_num(self, networkStr:list) -> list:
		Network = []
		for tmp in networkStr:
			Network.append(self.Can2num[self.ChannalMapping[tmp]])
		return Network

	#将起始字节位表示转换至起始位表示
	def bytebit2bit(self, byte:list, bit:list) -> list:
		ToBit = []
		for i in range(0,len(byte)):
			ToBit.append(int(byte[i]) * 8 + int(bit[i]))
		return ToBit

	#读取Excle一列数据
	def read_row(self, sheet, row:str) -> list:
		tmp_column_data = []
		for cell in list(sheet.columns)[column_index_from_string(row) - 1][2:]:
			if cell.value != None:
				tmp_column_data.append(cell.value)
		return tmp_column_data

	#读取所有列元素
	def read_rows_all(self, sheet):
		for i in range(1,column_index_from_string('V')):
			for j in range(1,3):
				#print(sheet[get_column_letter(i) + str(j)].value)
				if sheet[get_column_letter(i) + str(j)].value:
					if i <= 2:
						self.src_table[str(sheet[get_column_letter(i) + str(j)].value)] = read_row(sheet, get_column_letter(i))
					elif i <= column_index_from_string('L'):
						self.src_table["src_" + str(sheet[get_column_letter(i) + str(j)].value)] = read_row(sheet, get_column_letter(i))
					elif i <= column_index_from_string('T'):
						self.src_table["des_" + str(sheet[get_column_letter(i) + str(j)].value)]  = read_row(sheet, get_column_letter(i))
					else:
						self.src_table[str(sheet[get_column_letter(i) + str(j)].value)] = read_row(sheet, get_column_letter(i))

	#读取所有列元素
	def build_rows_all(self, SignalDatas):
		for i in range(column_index_from_string('U')):
			for j in range(0,2):
				#print(sheet[get_column_letter(i) + str(j)].value)
				if SignalDatas[i][j]:
					if i < 2:
						self.src_table[str(SignalDatas[i][j])] = SignalDatas[i][2:]
					elif i < column_index_from_string('L'):
						self.src_table["src_" + str(SignalDatas[i][j])] = SignalDatas[i][2:]
					elif i < column_index_from_string('T'):
						self.src_table["des_" + str(SignalDatas[i][j])]  = SignalDatas[i][2:]
					else:
						self.src_table[str(SignalDatas[i][j])] = SignalDatas[i][2:]

	#获取DTC标志
	def get_DTC(self) -> list:
		DTCTmp = []
		for i in range(len(self.src_table["信号名称"])):
			if self.src_table["DTC"][i] == "NA" or self.src_table["DTC"][i] == "None":
				DTCTmp.append('N')
			else:
				DTCTmp.append('Y')
		return DTCTmp

	#创建目标表列表
	def build_des_table(self) -> None:
		print(self.config.platInfo, "sig")
		if self.config.platInfo == "GAW1.2_OldPlatform" or self.config.platInfo == "GAW1.2_NewPlatform" or self.config.platInfo == "CHJ":
			self.des_table.append(self.src_table["信号名称"])
			self.des_table.append(self.get_TxMeesageName())
			self.des_table.append(self.src_table["des_目标网段ID"])
			self.des_table.append(self.src_table["des_周期"])
			self.des_table.append(self.src_table["des_dlc"])
			self.des_table.append(self.ToNetwork_num(self.src_table["des_目标网段"]))
			self.des_table.append(self.bytebit2bit(self.src_table["des_起始byte"], self.src_table["des_起始bit"]))
			self.des_table.append(self.src_table["des_信号长度"])
			self.des_table.append(self.get_RxMeesageName())
			self.des_table.append(self.src_table["src_源报文ID"])
			self.des_table.append(self.src_table["src_周期"])
			self.des_table.append(self.src_table["src_dlc"])
			self.des_table.append(self.ToNetwork_num(self.src_table["src_源网段"]))
			self.des_table.append(self.bytebit2bit(self.src_table["src_起始byte"], self.src_table["src_起始bit"]))
			self.des_table.append(self.src_table["src_信号长度"])
			self.des_table.append(self.src_table["src_信号格式"])
			self.des_table.append(self.get_DTC())
			self.des_table.append(self.src_table["src_初始值"])
			self.des_table.append(self.src_table["src_默认值"])
			self.des_table.append(self.src_table["des_目标网段"])
		elif self.config.platInfo == "Qoros_C6M0":
			self.des_table.append(self.src_table["信号名称"])
			self.des_table.append(self.get_TxMeesageName())
			self.des_table.append(self.src_table["des_目标网段ID"])
			self.des_table.append(self.src_table["des_周期"])
			self.des_table.append(self.src_table["des_dlc"])
			self.des_table.append(self.ToNetwork_num(self.src_table["des_目标网段"]))
			self.des_table.append(self.bytebit2bit(self.src_table["des_起始byte"], self.src_table["des_起始bit"]))
			self.des_table.append(self.src_table["des_信号长度"])
			self.des_table.append(self.src_table["des_信号格式"])
			self.des_table.append(self.get_RxMeesageName())
			self.des_table.append(self.src_table["src_源报文ID"])
			self.des_table.append(self.src_table["src_周期"])
			self.des_table.append(self.src_table["src_dlc"])
			self.des_table.append(self.ToNetwork_num(self.src_table["src_源网段"]))
			self.des_table.append(self.bytebit2bit(self.src_table["src_起始byte"], self.src_table["src_起始bit"]))
			self.des_table.append(self.src_table["src_信号长度"])
			self.des_table.append(self.src_table["src_信号格式"])
			self.des_table.append(self.get_DTC())
			self.des_table.append(self.src_table["src_初始值"])
			self.des_table.append(self.src_table["src_默认值"])
			self.des_table.append(self.src_table["des_目标网段"])


	# 信号列表处理，将列表中的所有int转为str
	def SignalDatas_handling(self, srcData:list) -> list:
		retDataList = []
		for i in range(len(srcData)):
			retDataList.append([str(j) for j in srcData[i]])

		return retDataList

	def mkdir(self, path:str): 
	    # 去除首位空格、尾部\
	    path=path.strip()
	    # 判断结果
	    if not os.path.exists(path):
	        os.makedirs(path) 

	
	# #主函数
	# def main(self):
	# 	if self.pathname != "":
	# 		# 默认可读写，若有需要可以指定write_only和read_only为True
	# 		wb = load_workbook(self.pathname)

	# 		# 根据sheet名字获得sheet
	# 		sheet = wb.get_sheet_by_name('Sheet2')

	# 		#获取通道映射
	# 		self.get_channalmapping(sheet)
	# 		#print(self.ChannalMapping)

	# 		#读取源表中的所有列存入字典中
	# 		self.read_rows_all(sheet)
	# 		print(self.src_table)
			
	# 		#创建目标表列表
	# 		self.build_des_table()

	# 		#将目标表数据写入CSV文件中
	# 		with open(self.pathname[:self.pathname.rfind('/')+1] + "SignalRoute.csv", 'w') as csvfile:	
	# 			writer = csv.writer(csvfile)
	# 			#写入文件头
	# 			writer.writerow(self.TableHeader)
	# 			#写入目标数据
	# 			writer.writerows(zip(*self.des_table))

		

	# 		print("------------Success------------------")
	# 	else:
	# 		print("请选择路由表")

	#主函数
	def main_pandas(self):
		if self.pathname != "": 
			# 读取指定列有效数据
			dataFrame = read_excel(self.pathname, sheet_name="Sheet2", header=None, na_values="", usecols="A:U")
			# print(dataFrame)
			# 将DataFrame格式数据转为列表
			SignalDatas = dataFrame.fillna("None").values.tolist()
			# zip迭代 + map映射
			SignalDatas = list(map(list, zip(*SignalDatas)))
			# 将列表中的所有int转为str
			SignalDatas = self.SignalDatas_handling(SignalDatas)

			# print(SignalDatas)

			self.build_rows_all(SignalDatas)

			# print(self.src_table)

			# 读取指定列数据
			dataFrame = read_excel(self.pathname, sheet_name="Sheet2", header=None, na_values="", usecols="V:X")
			# print(dataFrame)
			# 获取通道映射
			self.ChannalMapping = self.getChannalMapping_pandas(dataFrame, 0)
			# print(g_var.self.ChannalMapping)
			
			# #创建目标表列表
			self.build_des_table()

			self.mkdir("output")
			# 将目标表数据写入CSV文件中
			# with open(self.pathname[:self.pathname.rfind('/')+1] + "SignalRoute.csv", 'w', newline='') as csvfile:	
			with open("output/SignalRoute.csv", 'w', newline='') as csvfile:	
				writer = csv.writer(csvfile)
				#写入文件头
				writer.writerow(self.TableHeader)
				#写入目标数据
				writer.writerows(zip(*self.des_table))

			

			print("------------Success------------------")
		else:
			print("请选择路由表")

# #通道映射
# self.Can2num = {"CAN1":1, "CAN2":2, "CAN3":3, "CAN4":4, "CAN5":5, "CAN6":6}
# #目标Excel表头
# self.TableHeader = ["SignalName", "TxMeesageName", "TxCANID", "TxPeriod", "TxDLC", "TxChannle", \
# 				"TxStartBit", "TxSigLen", "RxMeesageName", "RxCANID", "RxPeriod", "RxDLC", \
# 				"RxChannel", "RxStartBit", "RxSigLen", "ByteOrder", "RxDTC", "inival", "dfVal", "desName"]
# self.ChannalMapping = {}
# self.src_table = {}
# self.des_table = []

if __name__ == '__main__':
	signal_table_convert = SignalTableConvert()
	signal_table_convert.get_file_pathname()
	signal_table_convert.main_pandas()

	# #读取指定路径中的Excel表
	# self.pathname = askopenfilename(filetypes = [("Excel",".xlsx")])

	# if self.pathname != '':
	# 	# main()
	# 	main_pandas()
