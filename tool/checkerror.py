import re
from copy import deepcopy
from openpyxl.utils import get_column_letter, column_index_from_string
from tkinter.filedialog import askopenfilename
from pandas import read_excel


class MsgCheckError(object):
	'''报文表错误检查'''
	def __init__(self):
		self.pathName = ""
		self.headerList = []
		self.dataList = []

	def get_file_pathname(self):
		"""获取路由表路径"""
		self.pathName = askopenfilename(filetypes = [("Excel",".xlsx")])
		if self.pathName == ():
			self.pathName = ""
			self.headerList = []
			self.dataList = []
		# print(self.pathName)
		# print(type(self.pathName))

	def read_data(self):
		"""从路由表中读取数据"""
		self.headerList = []
		self.dataList = []
		if self.pathName != "":
			dataFrame = read_excel(self.pathName, sheet_name="Sheet1", header=None, na_values="", usecols="A:Q")			
			self.dataList = dataFrame.fillna("None").values[2:].tolist()
			header = dataFrame.fillna("None").values[:2].tolist()

			for i in range(column_index_from_string("Q")):
				if i <= (column_index_from_string("F") - 1) or ((column_index_from_string("O") - 1) <= i <= (column_index_from_string("Q") - 1)):
					self.headerList.append(header[0][i])
				else:
					self.headerList.append(header[1][i])

		print(self.headerList)
		# print(self.dataList)

	def get_chmap(self):
		'''获取通道映射'''
		self.chMap = {}
		for i in range(6):
			self.chMap[self.dataList[i][self.headerList.index("CAN通道")]] = self.dataList[i][self.headerList.index("网段名称")]
		print(self.chMap)
		print(list(self.chMap.values()))
		# print(type(self.chMap.values()))
		print(list(self.chMap.keys()))

	def get_diag_ch(self):
		'''获取诊断通道'''
		self.diagch = []
		for i in range(18,20):
			self.diagch.append(self.dataList[i][column_index_from_string('N') - 1])
		print(self.diagch)

	def literal_check(self):
		'''字面语法性检查'''
		# 深拷贝一份路由数据，不破坏原有数据
		dataList = deepcopy(self.dataList)
		headerList = deepcopy(self.headerList)

		tab = "MsgTable"
		row = None # 对应于excel表的行
		col = None # 对应于excel表的列
		hint = "" # ""：没有错误

		ret = False

		# 对报文数据的每一列进行数据校验
		for i in range(len(dataList)):
			row = i + 3

			if i < 6:
				if not re.match("^CAN[1-6]{1}$", dataList[i][headerList.index("CAN通道")]):
					col = get_column_letter(headerList.index("CAN通道") + 1)
					hint = "CAN通道错误（请填写CAN1-CAN6)"
					break
				if not re.match("^[0-9 A-Z]+$", dataList[i][headerList.index("网段名称")], flags=re.IGNORECASE):
					col = get_column_letter(headerList.index("网段名称") + 1)
					hint = "网段名称错误（请填写字母和数字组合)"
					break
			if i == 18 or i == 19:
				if not re.match("^([0-9 A-Z]+)|None$", dataList[i][headerList.index("网段名称")], flags=re.IGNORECASE):
					col = get_column_letter(headerList.index("网段名称") + 1)
					hint = "网段名称错误（请填写字母和数字组合或NA)"
					break
				if dataList[i][headerList.index("网段名称")] != "None" and dataList[i][headerList.index("网段名称")] not in list(self.chMap.values()):
					col = get_column_letter(headerList.index("网段名称") + 1)
					hint = "网段名称与通道映射表中的网段名称不一致，请保持一致或填写NA"
					break

			if not re.match("^0x(([0-9 A-F]{1,2})|([0-7]{1}[0-9 A-Z]{2}))$", str(dataList[i][headerList.index("源报文ID")]), flags=re.IGNORECASE):
				col = get_column_letter(headerList.index("源报文ID") + 1)
				hint = "源报文ID格式错误或超出范围，请填写标准的十六进制格式(0x000-0x7FF))\n如为空则检查每一列是否行数一致"
				break
			if not re.match("^([1-9]{1}[0-9]+)|None$", str(dataList[i][headerList.index("周期")])):
				col = get_column_letter(headerList.index("周期") + 1)
				hint = "源报文周期错误（请填写十进制数字且不小于10或NA)"
				break
			if not re.match("^[1-8]{1}", str(dataList[i][headerList.index("dlc")])):
				col = get_column_letter(headerList.index("dlc") + 1)
				hint = "源报文长度错误，请填写1-8"
				break
			if not re.match("^([0-9 A-Z]{6})|None$", str(dataList[i][headerList.index("DTC码")]), flags=re.IGNORECASE):
				col = get_column_letter(headerList.index("DTC码") + 1)
				hint = "源报文DTC码错误，请填写正确格式或不填"
				break
			if not re.match("^([Y]{1})|None$", dataList[i][headerList.index("RxInterrupt")]):
				col = get_column_letter(headerList.index("RxInterrupt") + 1)
				hint = "源报文是否中断接收错误，请填写Y(大写)或不填"
				break
			if not re.match("^([0-9 A-Z]+)|None$", dataList[i][headerList.index("MsgNode")]):
				col = get_column_letter(headerList.index("MsgNode") + 1)
				hint = "源报文节点错误，请填写正确的节点名称或NA"
				break
			if not re.match("^[NY]{1}$", dataList[i][headerList.index("DiagMsg")]):
				col = get_column_letter(headerList.index("DiagMsg") + 1)
				hint = "源报文是否为诊断报文错误，请填写Y或N(大写)"
				break

		if hint:
			ret = (tab, row, col, hint)
		print(ret)
		return ret

	def logic_check(self):
		'''报文数据逻辑性检查'''
		# 深拷贝一份路由数据，不破坏原有数据
		dataList = deepcopy(self.dataList)
		headerList = deepcopy(self.headerList)

		id_ch_list = []
		first_msg_data_list = []

		tab = "MsgTable"
		row = None # 对应于excel表的行
		col = None # 对应于excel表的列
		hint = "" # ""：没有错误

		ret = False

		if len(list(set(list(self.chMap.keys())))) != 6:
			row = 3
			col = get_column_letter(headerList.index("CAN通道") + 1)
			hint = "通道映射表中CAN通道名称有重复，请依次填写CAN1->CAN6"
			return (tab, row, col, hint)

		if len(list(set(list(self.chMap.values())))) != 6:
			row = 3
			col = get_column_letter(headerList.index("网段名称") + 1)
			hint = "通道映射表中网段名称有重复，请勿重复；如通道未使用，请填写NCn，如通道2未使用，填NC2"
			return (tab, row, col, hint)

		if headerList[column_index_from_string("L") - 1] != self.chMap["CAN1"]:
			row = 2
			col = "L"
			hint = "L列目标网段名称与通道映射表中的CAN1对应的网段名称不一致，请填写一致"
			return (tab, row, col, hint)
		if headerList[column_index_from_string("K") - 1] != self.chMap["CAN2"]:
			row = 2
			col = "K"
			hint = "K列目标网段名称与通道映射表中的CAN2对应的网段名称不一致，请填写一致"
			return (tab, row, col, hint)
		if headerList[column_index_from_string("J") - 1] != self.chMap["CAN3"]:
			row = 2
			col = "J"
			hint = "J列目标网段名称与通道映射表中的CAN3对应的网段名称不一致，请填写一致"
			return (tab, row, col, hint)
		if headerList[column_index_from_string("I") - 1] != self.chMap["CAN4"]:
			row = 2
			col = "I"
			hint = "I列目标网段名称与通道映射表中的CAN4对应的网段名称不一致，请填写一致"
			return (tab, row, col, hint)
		if headerList[column_index_from_string("H") - 1] != self.chMap["CAN5"]:
			row = 2
			col = "H"
			hint = "H列目标网段名称与通道映射表中的CAN5对应的网段名称不一致，请填写一致"
			return (tab, row, col, hint)
		if headerList[column_index_from_string("G") - 1] != self.chMap["CAN6"]:
			row = 2
			col = "G"
			hint = "G列目标网段名称与通道映射表中的CAN6对应的网段名称不一致，请填写一致"
			return (tab, row, col, hint)

		if self.diagch[0] not in list(self.chMap.values()):
			row = 21
			col = "N"
			hint = "N列主诊断通道网段名称与通道映射表中的CAN6对应的网段名称不一致，请填写一致"
			return (tab, row, col, hint)
		if self.diagch[1] != "None" and self.diagch[1] not in list(self.chMap.values()):
			row = 22
			col = "N"
			hint = "N列次诊断通道网段名称与通道映射表中的CAN6对应的网段名称不一致，请填写一致或填写NA"
			return (tab, row, col, hint)

		for ind, subList in enumerate(dataList):
			row = ind +3
			# 进行单行报文数据校验
			if subList[headerList.index("源网段")] not in list(self.chMap.values()):
				col = get_column_letter(headerList.index("源网段") + 1)
				hint = "F列源网段名称与通道映射表中的网段名称不一致，请保持与N列一致"
				break
			# 前后一致性数据校验
			rx_id_ch = [subList[headerList.index("源报文ID")], subList[headerList.index("源网段")]]
			if rx_id_ch not in id_ch_list:
				id_ch_list.append(rx_id_ch)
				first_msg_data_list.append(subList)			
			else:
				first_msg_data = first_msg_data_list[id_ch_list.index(rx_id_ch)]
				# 逐列进行校验
				if subList[headerList.index("周期")] != first_msg_data[headerList.index("周期")]:
					col = get_column_letter(headerList.index("周期") + 1)
					hint = "源报文周期与该行之上同网段同ID源报文对应周期不一致，请保持一致"
					break
				if subList[msgHeaderList.index("dlc")] != first_msg_data[msgHeaderList.index("dlc")]:
					col = get_column_letter(msgHeaderList.index("dlc") + 1)
					hint = "源报文DLC与该行之上同网段同ID源报文对应DLC不一致，请保持一致"
					break
				if subList[msgHeaderList.index("DTC码")] != first_msg_data[msgHeaderList.index("DTC码")]:
					col = get_column_letter(msgHeaderList.index("DTC码") + 1)
					hint = "源报文DTC码与该行之上同网段同ID源报文对应DTC码不一致，请保持一致"
					break
				if subList[msgHeaderList.index("RxInterrupt")] != first_msg_data[msgHeaderList.index("RxInterrupt")]:
					col = get_column_letter(msgHeaderList.index("RxInterrupt") + 1)
					hint = "源报文RxInterrupt与该行之上同网段同ID源报文对应RxInterrupt不一致，请保持一致"
					break
				if subList[msgHeaderList.index("MsgNode")] != first_msg_data[msgHeaderList.index("MsgNode")]:
					col = get_column_letter(msgHeaderList.index("MsgNode") + 1)
					hint = "源报文MsgNode与该行之上同网段同ID源报文对应MsgNode不一致，请保持一致"
					break
				if subList[msgHeaderList.index("DiagMsg")] != first_msg_data[msgHeaderList.index("DiagMsg")]:
					col = get_column_letter(msgHeaderList.index("DiagMsg") + 1)
					hint = "源报文DiagMsg与该行之上同网段同ID源报文对应DiagMsg不一致，请保持一致"
					break
		if hint:
			ret = (tab, row, col, hint)
		print(ret)
		return ret

	def main(self):
		'''主函数'''
		self.read_data()
		self.get_chmap()
		self.get_diag_ch()
		ret = self.literal_check()
		if ret == False:
			ret = self.logic_check()

		return ret


class SignalCheckError(object):
	'''信号表错误检查'''
	def __init__(self):
		self.pathName = ""
		self.headerList = []
		self.dataList = []

	def get_file_pathname(self):
		"""获取路由表路径"""
		self.pathName = askopenfilename(filetypes = [("Excel",".xlsx")])
		if self.pathName == ():
			self.pathName = ""
			self.headerList = []
			self.dataList = []
		# print(self.pathName)
		# print(type(self.pathName))

	def read_data(self):
		"""从路由表中读取数据"""
		self.headerList = []
		self.dataList = []
		if self.pathName != "":
			dataFrame = read_excel(self.pathName, sheet_name="Sheet2", header=None, na_values="", usecols="A:X")			
			self.dataList = dataFrame.fillna("None").values[2:].tolist()
			header = dataFrame.fillna("None").values[:2].tolist()

			for i in range(column_index_from_string("X")):
				if i <= (column_index_from_string("H") - 1):
					if i <= (column_index_from_string("B") - 1):
						self.headerList.append(header[0][i])
					else:
						self.headerList.append("src_" + header[0][i])
				elif i <= (column_index_from_string("L") - 1):
					self.headerList.append("src_" + header[1][i])
				elif i <= (column_index_from_string("P") - 1):
					self.headerList.append("des_" + header[0][i])
				elif i <= (column_index_from_string("T") - 1):
					self.headerList.append("des_" + header[1][i])
				elif i <= (column_index_from_string("U") - 1):
					self.headerList.append(header[0][i])
				else:
					self.headerList.append(header[1][i])

		print(self.headerList)
		# print(self.dataList)

	def get_chmap(self):
		'''获取通道映射'''
		self.chMap = {}
		for i in range(6):
			self.chMap[self.dataList[i][self.headerList.index("CAN通道")]] = self.dataList[i][self.headerList.index("网段名称")]
		print(self.chMap)
		print(list(self.chMap.values()))
		# print(type(self.chMap.values()))
		print(list(self.chMap.keys()))

	def literal_check(self):
		'''字面语法性检查'''
		# 深拷贝一份路由数据，不破坏原有数据
		dataList = deepcopy(self.dataList)
		headerList = deepcopy(self.headerList)

		tab = "SignalTable"
		row = None # 对应于excel表的行
		col = None # 对应于excel表的列
		hint = "" # ""：没有错误

		ret = False

		# 对信号数据的每一列进行合法性校验
		for i in range(len(dataList)):
			row = i + 3
			if i < 6:
				if not re.match("^CAN[1-6]{1}$", dataList[i][headerList.index("CAN通道")]):
					col = get_column_letter(headerList.index("CAN通道") + 1)
					hint = "CAN通道错误（请填写CAN1->CAN6)"
					break
				if not re.match("^[0-9 A-Z]+$", dataList[i][headerList.index("网段名称")], flags=re.IGNORECASE):
					col = get_column_letter(headerList.index("网段名称") + 1)
					hint = "网段名称错误（请填写字母和数字组合)"
					break
				if not re.match("^NODE_[A-F]{1}$", dataList[i][headerList.index("程序节点")], flags=re.IGNORECASE):
					col = get_column_letter(headerList.index("程序节点") + 1)
					hint = "程序节点错误（请填写NODE_A->NODE_F)"
					break

			if not re.match("^0x(([0-9 A-F]{1,2})|([0-7]{1}[0-9 A-Z]{2}))$", str(dataList[i][headerList.index("src_源报文ID")]), flags=re.IGNORECASE):
				col = get_column_letter(headerList.index("src_源报文ID") + 1)
				hint = "源报文ID格式错误或超出范围，请填写标准的十六进制格式(0x000-0x7FF))\n如为空则检查每一列是否行数一致"
				break
			if not re.match("0x[0-9 A-F]+", str(dataList[i][headerList.index("src_默认值")]), flags=re.IGNORECASE):
				col = get_column_letter(headerList.index("src_默认值") + 1)
				hint = "源信号失效值填写错误，请填写标准的十六进制格式"
				break
			if not re.match("0x[0-9 A-F]+", str(dataList[i][headerList.index("src_初始值")]), flags=re.IGNORECASE):
				col = get_column_letter(headerList.index("src_初始值") + 1)
				hint = "源信号初始值填写错误，请填写标准的十六进制格式"
				break
			if not re.match("^[1-8]{1}$", str(dataList[i][headerList.index("src_dlc")])):
				col = get_column_letter(headerList.index("src_dlc") + 1)
				hint = "源信号DLC错误，请填写1-8"
				break
			if not re.match("^[1-9]{1}[0-9]+$", str(dataList[i][headerList.index("src_周期")])):
				col = get_column_letter(headerList.index("src_周期") + 1)
				hint = "源信号周期错误（填写十进制数字且大于10)"
				break
			if not (re.match("^[0-7]{1}$", str(dataList[i][headerList.index("src_起始byte")])) and (0 <= int(dataList[i][headerList.index("src_起始byte")]) <= 7)):
				col = get_column_letter(headerList.index("src_起始byte") + 1)
				hint = "源信号起始Byte错误，请填写0-7"
				break
			if not (re.match("^[0-7]{1}$", str(dataList[i][headerList.index("src_起始bit")])) and (0 <= int(dataList[i][headerList.index("src_起始bit")]) <= 7)):
				col = get_column_letter(headerList.index("src_起始bit") + 1)
				hint = "源信号起始Bit错误，请填写0-7"
				break
			if not (re.match("^[1-9]{1}[0-9]?$", str(dataList[i][headerList.index("src_信号长度")])) and (1 <= int(dataList[i][headerList.index("src_信号长度")]) <= 64)):
				col = get_column_letter(headerList.index("src_信号长度") + 1)
				hint = "源信号长度错误，请填写1-64"
				break
			if not re.match("^[0-1]{1}$", str(dataList[i][headerList.index("src_信号格式")])):
				col = get_column_letter(headerList.index("src_信号格式") + 1)
				hint = "源信号格式错误，请填写0或1，0表示Motorola大端，1表示Intel小端"
				break

			if not re.match("^0x(([0-9 A-F]{1,2})|([0-7]{1}[0-9 A-Z]{2}))$", str(dataList[i][headerList.index("des_目标网段ID")]), flags=re.IGNORECASE):
				col = get_column_letter(headerList.index("des_目标网段ID") + 1)
				hint = "目标报文ID格式错误或超出范围，请填写标准的十六进制格式(0x000-0x7FF))"
				break
			if not re.match("^[1-8]{1}$", str(dataList[i][headerList.index("des_dlc")])):
				col = get_column_letter(headerList.index("des_dlc") + 1)
				hint = "目标信号DLC错误，请填写1-8"
				break
			if not re.match("^[1-9]{1}[0-9]+$", str(dataList[i][headerList.index("des_周期")])):
				col = get_column_letter(headerList.index("des_周期") + 1)
				hint = "目标信号周期错误（填写十进制数字且大于10)"
				break
			if not (re.match("^[0-7]{1}$", str(dataList[i][headerList.index("des_起始byte")])) and (0 <= int(dataList[i][headerList.index("des_起始byte")]) <= 7)):
				col = get_column_letter(headerList.index("des_起始byte") + 1)
				hint = "目标信号起始Byte错误，请填写0-7"
				break
			if not (re.match("^[0-7]{1}$", str(dataList[i][headerList.index("des_起始bit")])) and (0 <= int(dataList[i][headerList.index("des_起始bit")]) <= 7)):
				col = get_column_letter(headerList.index("des_起始bit") + 1)
				hint = "目标信号起始Bit错误，请填写0-7"
				break
			if not (re.match("^[1-9]{1}[0-9]?$", str(dataList[i][headerList.index("des_信号长度")])) and (1 <= int(dataList[i][headerList.index("des_信号长度")]) <= 64)):
				col = get_column_letter(headerList.index("des_信号长度") + 1)
				hint = "目标信号长度错误，请填写1-64"
				break
			if not re.match("^[0-1]{1}$", str(dataList[i][headerList.index("des_信号格式")])):
				col = get_column_letter(headerList.index("des_信号格式") + 1)
				hint = "目标信号格式错误，请填写0或1，0表示Motorola大端，1表示Intel小端"
				break
			
			if not re.match("^([0-9 A-Z]{6})|None$", str(dataList[i][headerList.index("DTC")])):
				col = get_column_letter(headerList.index("DTC") + 1)
				hint = "源报文是否作为节点丢失DTC判断条件填写错误，请填写6位标准DTC码格式或填NA或不填"
				break
			

		if hint:
			ret = (tab, row, col, hint)
		print(ret)
		return ret

	def logic_check(self):
		'''报文数据逻辑性检查'''
		# 深拷贝一份路由数据，不破坏原有数据
		dataList = deepcopy(self.dataList)
		headerList = deepcopy(self.headerList)

		rx_id_ch_list = []
		first_rxsignal_data_list = []
		tx_id_ch_list = []
		first_txsignal_data_list = []

		tab = "SignalTable"
		row = None # 对应于excel表的行
		col = None # 对应于excel表的列
		hint = "" # ""：没有错误

		ret = False

		if len(list(set(list(self.chMap.keys())))) != 6:
			row = 3
			col = get_column_letter(headerList.index("CAN通道") + 1)
			hint = "通道映射表中CAN通道名称有重复，请依次填写CAN1->CAN6"
			return (tab, row, col, hint)

		if len(list(set(list(self.chMap.values())))) != 6:
			row = 3
			col = get_column_letter(headerList.index("网段名称") + 1)
			hint = "通道映射表中网段名称有重复，请勿重复；如通道未使用，请填写NCn，如通道2未使用，填NC2"
			return (tab, row, col, hint)


		# 将信号分为接收信号和发送信号分别处理；
		# 轮询每一个信号，对信号数据进行校验；
		# 提取ID和CH，如果ID和CH是第一次出现，将其放入列表中；
		# 如果ID和CH已经出现过，则将当前信号与已出现的信号数据进行校验；
		for ind, subList in enumerate(dataList):
			row = ind +3
			# 进行单行报文数据校验
			if subList[headerList.index("src_源网段")] not in list(self.chMap.values()):
				col = get_column_letter(headerList.index("src_源网段") + 1)
				hint = "H列源网段名称与通道映射表中的网段名称不一致，请保持一致"
				break
			if subList[headerList.index("des_目标网段")] not in list(self.chMap.values()):
				col = get_column_letter(headerList.index("des_目标网段") + 1)
				hint = "N列目标网段名称与通道映射表中的网段名称不一致，请保持一致"
				break

			# 接收信号数据校验			
			# 单行数据校验
			if str(subList[headerList.index("src_信号格式")]) == '0':
				if not (int(subList[headerList.index("src_信号长度")]) <= \
					int(subList[headerList.index("src_起始byte")]) * 8 + (8 - int(subList[headerList.index("src_起始bit")]))):
					col = get_column_letter(headerList.index("src_信号长度") + 1)
					hint = "源信号长度超限"
					break
			elif str(subList[headerList.index("src_信号格式")]) == '1':
				if not ((int(subList[headerList.index("src_起始byte")]) * 8 + int(subList[headerList.index("src_起始bit")]) + int(subList[headerList.index("RxSigLen")]) - 1) <= 63):
					col = get_column_letter(headerList.index("src_信号长度") + 1)
					hint = "源信号长度超限"
					break
			if not (int(subList[headerList.index("src_初始值")], 16) < 2**int(subList[headerList.index("src_信号长度")])):
				col = get_column_letter(headerList.index("src_初始值") + 1)
				hint = "源信号初始值超出合理值范围"
				break
			if not (int(subList[headerList.index("src_默认值")], 16) < 2**int(subList[headerList.index("src_信号长度")])):
				col = get_column_letter(headerList.index("src_默认值") + 1)
				hint = "源信号失效值超出合理值范围"
				break
			if not (int(subList[headerList.index("src_信号长度")]) == int(subList[headerList.index("des_信号长度")])):
				col = get_column_letter(headerList.index("src_信号长度") + 1)
				hint = "源信号长度与" + get_column_letter(headerList.index("des_信号长度") + 1) + "列目标信号长度不一致"
				break

			# 数据前后一致性校验
			rx_id_ch = [subList[headerList.index("src_源报文ID")], subList[headerList.index("src_源网段")]]
			if rx_id_ch not in rx_id_ch_list:
				rx_id_ch_list.append(rx_id_ch)
				first_rxsignal_data_list.append(subList)
			else:
				# 逐列进行校验
				first_rxsignal_data = first_rxsignal_data_list[rx_id_ch_list.index(rx_id_ch)]
				if subList[headerList.index("src_周期")] != first_rxsignal_data[headerList.index("src_周期")]:
					col = get_column_letter(headerList.index("src_周期") + 1)
					hint = "源信号周期与该行之上同网段同ID信号周期不一致"
					break
				if subList[headerList.index("src_dlc")] != first_rxsignal_data[headerList.index("src_dlc")]:
					col = get_column_letter(headerList.index("src_dlc") + 1)
					hint = "源信号DLC与该行之上同网段同ID信号DLC不一致"
					break
				if subList[headerList.index("src_信号格式")] != first_rxsignal_data[headerList.index("src_信号格式")]:
					col = get_column_letter(headerList.index("src_信号格式") + 1)
					hint = "源信号字节序与该行之上同网段同ID信号字节序不一致"
					break
				if subList[headerList.index("DTC")] != first_rxsignal_data[headerList.index("DTC")]:
					col = get_column_letter(headerList.index("RxDTC") + 1)
					hint = "源信号节点丢失DTC配置与该行之上同网段同ID信号节点丢失DTC配置不一致"
					break
			# 发送信号数据校验
			# 单行数据校验
			if str(subList[headerList.index("des_信号格式")]) == '0':
				if not (int(subList[headerList.index("des_信号长度")]) <= \
					int(subList[headerList.index("des_起始byte")]) * 8 + (8 - int(subList[headerList.index("des_起始bit")]))):
					col = get_column_letter(headerList.index("des_信号长度") + 1)
					hint = "目标信号长度超限"
					break
			elif str(subList[headerList.index("des_信号格式")]) == '1':
				if not ((int(subList[headerList.index("des_起始byte")]) * 8 + int(subList[headerList.index("des_起始bit")]) + int(subList[headerList.index("des_信号长度")]) - 1) <= 63):
					col = get_column_letter(headerList.index("des_信号长度") + 1)
					hint = "目标信号长度超限"
					break
			if not (int(subList[headerList.index("des_信号长度")]) == int(subList[headerList.index("src_信号长度")])):
				col = get_column_letter(headerList.index("des_信号长度") + 1)
				hint = "目标信号长度与接收信号长度不一致"
				break

			tx_id_ch = [subList[headerList.index("des_目标网段ID")], subList[headerList.index("des_目标网段")]]	
			if tx_id_ch not in tx_id_ch_list:
				tx_id_ch_list.append(tx_id_ch)
				first_txsignal_data_list.append(subList)

			else:
				# 逐列进行数据校验
				first_txsignal_data = first_txsignal_data_list[tx_id_ch_list.index(tx_id_ch)]
				if subList[headerList.index("des_周期")] != first_txsignal_data[headerList.index("des_周期")]:
					col = get_column_letter(headerList.index("des_周期") + 1)
					hint = "目标信号周期与该行之上同网段同ID信号周期不一致"
					break
				if subList[headerList.index("des_dlc")] != first_txsignal_data[headerList.index("des_dlc")]:
					col = get_column_letter(headerList.index("des_dlc") + 1)
					hint = "目标信号DLC与该行之上同网段同ID信号DLC不一致"
					break
				if subList[headerList.index("des_信号格式")] != first_txsignal_data[headerList.index("des_信号格式")]:
					col = get_column_letter(headerList.index("des_信号格式") + 1)
					hint = "目标信号字节序与该行之上同网段同ID信号字节序不一致"
					break

		if hint:
			ret = (tab, row, col, hint)
		print(ret)

		return ret

	def main(self):
		'''主函数'''
		self.read_data()
		self.get_chmap()
		ret = self.literal_check()
		if ret == False:
			ret = self.logic_check()
		return ret

		print("------------------Finished--------------------------")


if __name__ == '__main__':
	# msgCheckError = MsgCheckError()
	# msgCheckError.get_file_pathname()
	# msgCheckError.read_data()
	# msgCheckError.get_chmap()
	# msgCheckError.get_diag_ch()
	# msgCheckError.literal_check()
	# msgCheckError.logic_check()

	signalCheckError = SignalCheckError()
	signalCheckError.get_file_pathname()
	signalCheckError.read_data()
	signalCheckError.get_chmap()
	signalCheckError.literal_check()
	signalCheckError.logic_check()

