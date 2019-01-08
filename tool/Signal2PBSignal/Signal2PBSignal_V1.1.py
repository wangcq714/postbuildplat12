import pandas

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
import csv
from tkinter import *
from tkinter.filedialog import askdirectory,askopenfilename

#获取通道映射
def get_channalmapping(sheet):
	for i in range(6):
		ChannalMapping[sheet['W' + str(i + 3)].value] = sheet['V' + str(i + 3)].value

#通道映射,如果dir==0,则为CAN1：netname,dir == 1, netname:dir
def getChannalMapping_pandas(dataFrame, dir) -> dict:
	ChannalMapping = {}
	if dir == 0:
		for i in range(6):
			# ChannalMapping[dataFrame.values[i + 2][column_index_from_string('N') - 1]] = dataFrame.values[i +2][column_index_from_string('M') - 1].lower()
			ChannalMapping[dataFrame.values[i + 2][1]] = dataFrame.values[i +2][0]
	elif dir == 1:
		for i in range(6):
			# ChannalMapping[dataFrame.values[i + 2][column_index_from_string('M') - 1]] = dataFrame.values[i +2][column_index_from_string('N') - 1]
			ChannalMapping[dataFrame.values[i + 2][0]] = dataFrame.values[i +2][1]
	return ChannalMapping

#获取发送报文的名字
def get_TxMeesageName() -> list:
	TxMeesageName = []
	for i in range(len(src_table["des_目标网段ID"])):
		TxMeesageName.append(src_table["des_目标网段"][i] + '_' + src_table["des_目标网段ID"][i][2:]) 
	return TxMeesageName

#获取接收报文的名字
def get_RxMeesageName() -> list:
	RxMeesageName = []
	for i in range(len(src_table["src_源报文ID"])):
		RxMeesageName.append("GW_" + src_table["src_源报文ID"][i][2:]) 
	return RxMeesageName

#获取发送至目标网段的数字代号
def ToNetwork_num(networkStr:list) -> list:
	Network = []
	for tmp in networkStr:
		Network.append(Can2num[ChannalMapping[tmp]])
	return Network

#将起始字节位表示转换至起始位表示
def bytebit2bit(byte:list, bit:list) -> list:
	ToBit = []
	for i in range(0,len(byte)):
		ToBit.append(int(byte[i]) * 8 + int(bit[i]))
	return ToBit

#读取Excle一列数据
def read_row(sheet, row:str) -> list:
	tmp_column_data = []
	for cell in list(sheet.columns)[column_index_from_string(row) - 1][2:]:
		if cell.value != None:
			tmp_column_data.append(cell.value)
	return tmp_column_data

#读取所有列元素
def read_rows_all(sheet):
	for i in range(1,column_index_from_string('V')):
		for j in range(1,3):
			#print(sheet[get_column_letter(i) + str(j)].value)
			if sheet[get_column_letter(i) + str(j)].value:
				if i <= 2:
					src_table[str(sheet[get_column_letter(i) + str(j)].value)] = read_row(sheet, get_column_letter(i))
				elif i <= column_index_from_string('L'):
					src_table["src_" + str(sheet[get_column_letter(i) + str(j)].value)] = read_row(sheet, get_column_letter(i))
				elif i <= column_index_from_string('T'):
					src_table["des_" + str(sheet[get_column_letter(i) + str(j)].value)]  = read_row(sheet, get_column_letter(i))
				else:
					src_table[str(sheet[get_column_letter(i) + str(j)].value)] = read_row(sheet, get_column_letter(i))

#读取所有列元素
def build_rows_all(SignalDatas):
	for i in range(column_index_from_string('U')):
		for j in range(0,2):
			#print(sheet[get_column_letter(i) + str(j)].value)
			if SignalDatas[i][j]:
				if i < 2:
					src_table[str(SignalDatas[i][j])] = SignalDatas[i][2:]
				elif i < column_index_from_string('L'):
					src_table["src_" + str(SignalDatas[i][j])] = SignalDatas[i][2:]
				elif i < column_index_from_string('T'):
					src_table["des_" + str(SignalDatas[i][j])]  = SignalDatas[i][2:]
				else:
					src_table[str(SignalDatas[i][j])] = SignalDatas[i][2:]

#获取DTC标志
def get_DTC() -> list:
	DTCTmp = []
	for i in range(len(src_table["信号名称"])):
		if src_table["DTC"][i] == "NA" or src_table["DTC"][i] == "None":
			DTCTmp.append('N')
		else:
			DTCTmp.append('Y')
	return DTCTmp

#创建目标表列表
def build_des_table() -> None:
	des_table.append(src_table["信号名称"])
	des_table.append(get_TxMeesageName())
	des_table.append(src_table["des_目标网段ID"])
	des_table.append(src_table["des_周期"])
	des_table.append(src_table["des_dlc"])
	des_table.append(ToNetwork_num(src_table["des_目标网段"]))
	des_table.append(bytebit2bit(src_table["des_起始byte"], src_table["des_起始bit"]))
	des_table.append(src_table["des_信号长度"])
	des_table.append(get_RxMeesageName())
	des_table.append(src_table["src_源报文ID"])
	des_table.append(src_table["src_周期"])
	des_table.append(src_table["src_dlc"])
	des_table.append(ToNetwork_num(src_table["src_源网段"]))
	des_table.append(bytebit2bit(src_table["src_起始byte"], src_table["src_起始bit"]))
	des_table.append(src_table["src_信号长度"])
	des_table.append(src_table["des_信号格式"])
	des_table.append(get_DTC())
	des_table.append(src_table["src_初始值"])
	des_table.append(src_table["src_默认值"])
	des_table.append(src_table["des_目标网段"])


# 信号列表处理，将列表中的所有int转为str
def SignalDatas_handling(srcData:list) -> list:
	retDataList = []
	for i in range(len(srcData)):
		retDataList.append([str(j) for j in srcData[i]])

	return retDataList

	
#主函数
def main():
	# 默认可读写，若有需要可以指定write_only和read_only为True
	wb = load_workbook(pathname)

	# 根据sheet名字获得sheet
	sheet = wb.get_sheet_by_name('Sheet2')

	#获取通道映射
	get_channalmapping(sheet)
	#print(ChannalMapping)

	#读取源表中的所有列存入字典中
	read_rows_all(sheet)
	print(src_table)
	
	#创建目标表列表
	build_des_table()

	#将目标表数据写入CSV文件中
	with open(pathname[:pathname.rfind('/')+1] + "SignalRoute.csv", 'w') as csvfile:	
		writer = csv.writer(csvfile)
		#写入文件头
		writer.writerow(TableHeader)
		#写入目标数据
		writer.writerows(zip(*des_table))

	

	print("------------Success------------------")

#主函数
def main_pandas():
	global ChannalMapping

	# 读取指定列有效数据
	dataFrame = pandas.read_excel(pathname, sheet_name="Sheet2", header=None, na_values="", usecols="A:U")
	# print(dataFrame)
	# 将DataFrame格式数据转为列表
	SignalDatas = dataFrame.fillna("None").values.tolist()
	# zip迭代 + map映射
	SignalDatas = list(map(list, zip(*SignalDatas)))
	# 将列表中的所有int转为str
	SignalDatas = SignalDatas_handling(SignalDatas)

	# print(SignalDatas)

	build_rows_all(SignalDatas)

	# print(src_table)

	# 读取指定列数据
	dataFrame = pandas.read_excel(pathname, sheet_name="Sheet2", header=None, na_values="", usecols="V:X")
	# print(dataFrame)
	# 获取通道映射
	ChannalMapping = getChannalMapping_pandas(dataFrame, 0)
	# print(g_var.ChannalMapping)
	
	# #创建目标表列表
	build_des_table()

	# 将目标表数据写入CSV文件中
	with open(pathname[:pathname.rfind('/')+1] + "SignalRoute.csv", 'w') as csvfile:	
		writer = csv.writer(csvfile)
		#写入文件头
		writer.writerow(TableHeader)
		#写入目标数据
		writer.writerows(zip(*des_table))

	

	print("------------Success------------------")

#通道映射
Can2num = {"CAN1":1, "CAN2":2, "CAN3":3, "CAN4":4, "CAN5":5, "CAN6":6}
#目标Excel表头
TableHeader = ["SignalName", "TxMeesageName", "TxCANID", "TxPeriod", "TxDLC", "TxChannle", \
				"TxStartBit", "TxSigLen", "RxMeesageName", "RxCANID", "RxPeriod", "RxDLC", \
				"RxChannel", "RxStartBit", "RxSigLen", "ByteOrder", "RxDTC", "inival", "dfVal", "desName"]
ChannalMapping = {}
src_table = {}
des_table = []

#读取指定路径中的Excel表
pathname = askopenfilename(filetypes = [("Excel",".xlsx")])

if pathname != '':
	# main()
	main_pandas()
