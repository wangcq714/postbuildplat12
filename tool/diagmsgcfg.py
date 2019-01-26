import sys
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
import csv
from tkinter.filedialog import askopenfilename
import pandas


class DiagResTable(object):
	def __init__(self):
		self.Can2num = {"CAN1":1, "CAN2":2, "CAN3":3, "CAN4":4, "CAN5":5, "CAN6":6}
		self.pathname = ""
		self.ChannalMapping = {}
		self.validDataList = []
		self.resDataRxList = []
		self.resDataTxList = []
		self.reqDataList = []
		self.resMsgCfgList = []
		self.reqMsgCfgList = []
		self.diagch = []

	# 获取文件路径名
	def get_file_pathname(self):
		# 选择文件
		self.pathname = askopenfilename(filetypes = [("Excel",".xlsx")])
		if self.pathname == '':
			print("没有选择文件")
			exit()

	# 读取Excel
	def read_data(self):
		dataFrame = pandas.read_excel(self.pathname, sheet_name="Sheet1", header=None, na_values="", usecols="A:Q")
		self.msgDataList = dataFrame.fillna("None").values[2:].tolist()
		# print(self.msgDataList)
		self.msgDataList = self.int2str(self.msgDataList)
		# print(self.msgDataList)

	#获取通道映射
	def get_channalmapping_pandas(self):
		for i in range(6):
			self.ChannalMapping[self.msgDataList[i][column_index_from_string('N') - 1]] = self.msgDataList[i][column_index_from_string('M') - 1]
		print(self.ChannalMapping)

	#获取诊断通道
	def get_diag_ch(self):
		for i in range(18,20):
			self.diagch.append(self.msgDataList[i][column_index_from_string('N') - 1])
		print(self.diagch)

	# 获取有效数据
	def get_valid_data(self):
		for subList in self.msgDataList:
			if subList[column_index_from_string('Q') - 1].strip(' ') == 'Y':
				self.validDataList.append(subList)
		# print(self.validDataList)

	#获取要发送的通道
	def get_tx_ch(self, dataList) -> int:
		ch = 1
		for tick in dataList[column_index_from_string('L') - 1 : column_index_from_string('G') -2 : -1]:
			if tick == '√':
				break
			ch += 1

		return ch

	# 诊断响应报文接收MO配置
	def __res_msg_rx_element(self, dataList):
		retDataList = []
		retDataList.append(dataList[column_index_from_string('A') - 1].strip(' '))
		retDataList.append("NODE_" + get_column_letter(self.Can2num[self.ChannalMapping[dataList[column_index_from_string('F') - 1].strip(' ')]]))
		retDataList.append("MSG_RX")
		# retDataList.append(get_column_letter(self.Can2num[self.ChannalMapping[dataList[column_index_from_string('F') - 1].strip(' ')]]) + "_RR_" + \
		# 					dataList[column_index_from_string('B') - 1].strip(' ')[:dataList[column_index_from_string('B') - 1].strip(' ').rfind('_')])
		retDataList.append(get_column_letter(self.Can2num[self.ChannalMapping[dataList[column_index_from_string('F') - 1].strip(' ')]]) + "_RR_" + \
							dataList[column_index_from_string('B') - 1].strip(' '))
		retDataList.append(' '*(10 - len(dataList[column_index_from_string('B') - 1].strip(' ')[:dataList[column_index_from_string('B') - 1].strip(' ').rfind('_')])) + "0x7FFUL")
		retDataList.append("CAN_STD")
		retDataList.append(dataList[column_index_from_string('D') - 1].strip(' '))
		retDataList.append("FALSE")
		retDataList.append("0xFF")
		retDataList.append("NULL")

		return retDataList

	# 诊断响应报文发送MO配置
	def __res_msg_tx_element(self, dataList):
		retDataList = []
		retDataList.append(dataList[column_index_from_string('A') - 1].strip(' '))
		retDataList.append("NODE_A")
		retDataList.append("MSG_TX")
		# retDataList.append("A_RS_" + dataList[column_index_from_string('B') - 1].strip(' ')[:dataList[column_index_from_string('B') - 1].strip(' ').rfind('_')])
		retDataList.append("A_RS_" + dataList[column_index_from_string('B') - 1].strip(' '))
		retDataList.append(' '*(10 - len(dataList[column_index_from_string('B') - 1].strip(' ')[:dataList[column_index_from_string('B') - 1].strip(' ').rfind('_')])) + "0x7FFUL")
		retDataList.append("CAN_STD")
		retDataList.append(dataList[column_index_from_string('D') - 1].strip(' '))
		retDataList.append("FALSE")
		retDataList.append("0xFF")
		retDataList.append("NULL")

		return retDataList

	# 诊断请求报文配置表元素
	def __req_msg_element(self, dataList) -> list:
		retDataList = []
		# retDataList.append("DiagBasic_Rx_" + dataList[column_index_from_string('B') - 1].strip(' ')[:dataList[column_index_from_string('B') - 1].strip(' ').rfind('_')])
		retDataList.append("DiagBasic_Rx_" + dataList[column_index_from_string('B') - 1].strip(' '))
		retDataList.append(' '*(10 - len(dataList[column_index_from_string('B') - 1].strip(' ')[:dataList[column_index_from_string('B') - 1].strip(' ').rfind('_')])) + get_column_letter(self.get_tx_ch(dataList)) + "_RS_PHY")
		retDataList.append(dataList[column_index_from_string('A') - 1].strip(' '))

		return retDataList

	# 对得到的有效数据进行加工处理
	def data_handling(self):
		for subList in self.validDataList:
			if subList[column_index_from_string('F') - 1].strip(' ') != self.diagch[0] and subList[column_index_from_string('F') - 1].strip(' ') != self.diagch[1]:
			# if subList[column_index_from_string('F') - 1].strip(' ') != sys.argv[1]:
				self.resDataRxList.append(self.__res_msg_rx_element(subList))
				self.resDataTxList.append(self.__res_msg_tx_element(subList))
			else:
				if subList[column_index_from_string('A') - 1].strip(' ') != "0x7DF":
					self.reqDataList.append(self.__req_msg_element(subList))

		# print(self.resDataRxList)
		# print(self.resDataTxList)
		# print(self.reqDataList)

	# 创建诊断响应报文配置表
	def buid_res_msg_list(self) -> list:
		self.resMsgCfgList.append('{\n')
		for subList in self.resDataRxList:
			self.resMsgCfgList.append('\t' + "{")
			self.resMsgCfgList.append(', '.join(subList))
			self.resMsgCfgList.append("},\n")
		self.resMsgCfgList.append("\n")	
		for subList in self.resDataTxList:
			self.resMsgCfgList.append('\t' + "{")
			self.resMsgCfgList.append(', '.join(subList))
			self.resMsgCfgList.append("},\n")
		self.resMsgCfgList.append('}\n')

	# 创建诊断请求报文配置表
	def buid_req_msg_list(self) -> list:
		self.reqMsgCfgList.append("{\n")
		for subList in self.reqDataList:
			self.reqMsgCfgList.append('\t' + "{")
			self.reqMsgCfgList.append(', '.join(subList))
			self.reqMsgCfgList.append("},\n")
		self.reqMsgCfgList.append("}")

	def mkdir(self, path:str): 
	    # 去除首位空格、尾部\
	    path=path.strip()
	    # 判断结果
	    if not os.path.exists(path):
	        os.makedirs(path) 

	# 写入.c文件
	def write2file(self):
		self.mkdir("output")
		# with open(self.pathname[:self.pathname.rfind('/')+1] + "DiagCfgTable.c",'w') as Cfile:
		with open("output/DiagCfgTable.c",'w') as Cfile:
			#将FireWallTableList数组写入
			for tmp in self.resMsgCfgList:
				Cfile.write(tmp)
			Cfile.write("\n\n")
			for tmp in self.reqMsgCfgList:
				Cfile.write(tmp)

	# 写入.c文件
	def write2diagreqfile(self):
		self.mkdir("output")
		# with open(self.pathname[:self.pathname.rfind('/')+1] + "DiagReqCfgTable.c",'w') as Cfile:
		with open("output/DiagReqCfgTable.c",'w') as Cfile:
			#将FireWallTableList数组写入
			for tmp in self.reqMsgCfgList:
				Cfile.write(tmp)

	# 写入.c文件
	def write2diagresfile(self):
		self.mkdir("output")
		# with open(self.pathname[:self.pathname.rfind('/')+1] + "DiagResCfgTable.c",'w') as Cfile:
		with open("output/DiagResCfgTable.c",'w') as Cfile:
			#将FireWallTableList数组写入
			for tmp in self.resMsgCfgList:
				Cfile.write(tmp)

	# 将列表中的所有int转为str
	@staticmethod
	def int2str(dataList) -> list:
		retDataList = []
		for i in range(len(dataList)):
			retDataList.append([str(j) for j in dataList[i]])

		return retDataList


	def main(self):
		#判断程序是否传参
		# try:
		# 	Num = sys.argv[1]
		# except (IndexError,):
		# 	print("请给程序传递一个参数:路由表中诊断CAN名字！！！")
		# 	exit()
			
		# diagResTable = DiagResTable()

		self.get_file_pathname()

		self.read_data()

		self.get_channalmapping_pandas()
		self.get_diag_ch()

		self.get_valid_data()

		self.data_handling()

		self.buid_res_msg_list()

		self.buid_req_msg_list()

		self.write2file()

		print("------------------Finished--------------------------")

	def main_req(self):
		#判断程序是否传参
		# try:
		# 	Num = sys.argv[1]
		# except (IndexError,):
		# 	print("请给程序传递一个参数:路由表中诊断CAN名字！！！")
		# 	exit()
			
		# diagResTable = DiagResTable()

		# self.get_file_pathname()

		self.read_data()

		self.get_channalmapping_pandas()
		self.get_diag_ch()

		self.get_valid_data()

		self.data_handling()

		self.buid_res_msg_list()

		self.buid_req_msg_list()

		self.write2diagreqfile()

		print("------------------Finished--------------------------")

	def main_res(self):
		#判断程序是否传参
		# try:
		# 	Num = sys.argv[1]
		# except (IndexError,):
		# 	print("请给程序传递一个参数:路由表中诊断CAN名字！！！")
		# 	exit()
			
		# diagResTable = DiagResTable()

		# self.get_file_pathname()

		self.read_data()

		self.get_channalmapping_pandas()
		self.get_diag_ch()

		self.get_valid_data()

		self.data_handling()

		self.buid_res_msg_list()

		self.buid_req_msg_list()

		self.write2diagresfile()

		print("------------------Finished--------------------------")		


if __name__ == '__main__':
	diagResTable = DiagResTable()
	diagResTable.main()

